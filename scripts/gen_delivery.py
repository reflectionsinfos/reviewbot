import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DELIVERY_ITEMS = [
    ("1.1","Scope, Planning & Governance","Are scope / SoW / baselines clearly defined, signed off, and tracked?","Signed SoW or Statement of Work in project repository or document management system; scope baseline in PM tool (Jira/Azure DevOps) with change log showing approved changes"),
    ("1.2","Scope, Planning & Governance","Are change requests, assumptions, and dependencies captured and actively managed?","Change Request log with status, owner, and decision date; Assumptions & Dependency register linked from project plan; no unreviewed changes older than 5 business days"),
    ("1.3","Scope, Planning & Governance","Is the project plan (milestones, releases, critical path) realistic, updated, and communicated?","Project plan (MS Project/Jira/Azure DevOps) with milestones and critical path visible; last-updated within 2 weeks; plan shared with stakeholders via email or Confluence"),
    ("1.4","Scope, Planning & Governance","Is governance cadence followed (status reviews, steering committee, reporting)?","Meeting minutes or status reports for last 4+ weeks in shared folder or Confluence; steering committee deck from last meeting; recurring calendar invites in place"),
    ("1.5","Scope, Planning & Governance","Are decision logs and action items tracked with owners and due dates?","RAID log or Decision Register with owner, due date, and status columns; action item tracker updated within last 7 days; no overdue actions without escalation"),
    ("2.1","Delivery Health","Are milestones, sprints, and releases on track against plan (schedule variance within tolerance)?","Sprint burndown or velocity chart in Jira/Azure DevOps; milestone tracker with % complete vs plan; schedule variance within +/-10%; release forecast updated in project plan"),
    ("2.2","Delivery Health","Are deliverables peer-reviewed and meeting acceptance / Definition of Done criteria?","Definition of Done checklist visible in Jira or repo wiki; PR history showing peer reviews before merge; customer acceptance sign-off emails or comments on recent deliverables"),
    ("2.3","Delivery Health","Are blockers, technical debt, and rework under control and visible in plans?","Blocker or impediment board in Jira/Azure DevOps updated this sprint; technical debt backlog items with story points and priority; rework items tagged and quantified in current sprint"),
    ("2.4","Delivery Health","Is release planning (scope, dates, dependencies) clear and agreed with stakeholders?","Release plan document or Confluence page listing scope, go-live date, and dependencies; stakeholder sign-off email or meeting minutes confirming agreement"),
    ("3.1","Requirements & Customer Alignment","Are requirements / user stories clear, prioritized, approved, and version-controlled?","Backlog in Jira/Azure DevOps with acceptance criteria on all stories in current and next sprint; BRD or FRD document versioned in Git or Confluence with approval signatures"),
    ("3.2","Requirements & Customer Alignment","Are acceptance criteria and non-functional requirements defined and agreed with the customer?","Acceptance criteria on each user story in backlog; NFR document listing performance, security, and reliability targets with customer sign-off; NFR traceability to test cases"),
    ("3.3","Requirements & Customer Alignment","Are backlog refinement, demos, and walk-throughs regularly done with key stakeholders?","Sprint review or demo recordings/slides/meeting minutes from last 3 sprints; backlog refinement session notes; stakeholder attendance list showing business participation"),
    ("3.4","Requirements & Customer Alignment","Are stakeholder communications timely, transparent, and documented?","Status report emails or Confluence updates published on agreed cadence for last 4 weeks; communication plan document; distribution list with all key stakeholders confirmed"),
    ("3.5","Requirements & Customer Alignment","Are there any customer escalations, dissatisfaction signals, or relationship risks?","Escalation register or risk log entries tagged 'customer relationship'; CSAT scores or qualitative feedback from last QBR; resolution actions documented with owner and target date if escalations exist"),
    ("4.1","Risks, Issues & Escalations","Are project risks, issues, assumptions, and dependencies logged, owned, and regularly updated?","RAID log in Confluence/SharePoint with owner, probability, impact, mitigation, and last-reviewed date within 2 weeks; no entries older than 4 weeks without a review update"),
    ("4.2","Risks, Issues & Escalations","Are mitigations and contingency plans defined for high / critical risks?","RAID log showing mitigation and contingency columns populated for all Red/Amber risks; contingency plans reviewed by project sponsor; risk response plans communicated to relevant owners"),
    ("4.3","Risks, Issues & Escalations","Are issues resolved within agreed SLA / timelines?","Issue tracker showing resolution times vs SLA targets; no open P1/P2 issues beyond SLA; escalation evidence for overdue items; issue closure rate chart for last 3 sprints"),
    ("4.4","Risks, Issues & Escalations","Is the escalation matrix known, used effectively, and working as intended?","Escalation matrix document shared with all team members and customer contacts; evidence of escalation being used appropriately in last quarter; no stale unresolved escalations"),
    ("4.5","Risks, Issues & Escalations","Are repeated risks / issues feeding back into planning and process improvements?","Retrospective action items tracking recurring issues; lessons-learned log with improvements applied to current plan; recurring risk trend visible in RAID log history"),
    ("5.1","Resource & Team Management","Do we have the right mix of skills, capacity, and seniority to meet commitments?","Resource plan with skills matrix, allocation %, and seniority levels mapped to deliverables; no critical skill gaps unaddressed in RAID log; capacity sufficient for committed sprint scope"),
    ("5.2","Resource & Team Management","Are there any attrition, availability, or productivity gaps impacting delivery?","Resource risk entries in RAID log with mitigation; headcount tracker showing planned vs actual allocation; Jira velocity data showing stable or improving trend over last 3 sprints"),
    ("5.3","Resource & Team Management","Is there backup / succession for critical roles and single points of failure?","Knowledge transfer documentation for key roles; backup person identified in resource plan for each critical role; cross-training evidence or pair-programming records in capability matrix"),
    ("5.4","Resource & Team Management","Are onboarding, knowledge transfer, and role clarity adequate for new team members?","Onboarding checklist completed for all joiners in last quarter; RACI document showing clear role/responsibility boundaries; developer onboarding guide in repo wiki or Confluence"),
    ("5.5","Resource & Team Management","How is team morale and collaboration (within the team and with the customer)?","Team health check survey results or retrospective sentiment scores from last 3 sprints; absence/leave patterns within normal range; no unresolved interpersonal issues in retrospective actions"),
    ("6.1","Quality & Testing","Are unit, integration, and system tests executed with required coverage and documented results?","CI pipeline test reports showing pass rate and coverage at or above defined threshold; test execution report in XRAY/TestRail/Zephyr; defect density within agreed threshold"),
    ("6.2","Quality & Testing","Is UAT planned and executed with low defect leakage into production?","UAT test plan signed by business owner; UAT sign-off document or email from customer; post-UAT defect leakage metric below agreed threshold (e.g., < 5% of P1/P2 defects reaching production)"),
    ("6.3","Quality & Testing","Is test automation (functional / regression / smoke) leveraged appropriately?","Automated test suite in repository (Selenium/Cypress/Playwright/JUnit/pytest); CI pipeline config running automation on every build or PR; automation coverage report showing % of regression tests automated"),
    ("6.4","Quality & Testing","Are code reviews, static analysis, and non-functional tests (performance, security, reliability) in place?","Branch protection requiring PR review before merge; SonarQube/Checkmarx scan report with no unresolved critical issues; performance and security test results measured against NFR thresholds"),
    ("6.5","Quality & Testing","Are defect trends monitored and corrective actions identified for recurring issues?","Defect trend chart in Jira/Azure DevOps showing open/closed/new per sprint; defect age report; RCA documented for all critical defects; recurring defect categories tracked and actioned"),
    ("7.1","Compliance & Security","Are organisational processes followed (SQA, documentation, approvals, audits)?","SQA checklist completed for current delivery phase; document approval records with sign-off timestamps in Confluence/SharePoint; internal audit findings log with closure evidence"),
    ("7.2","Compliance & Security","Is customer data handled securely (access control, encryption, environments, masking)?","Data handling policy signed by all team members; access control matrix showing least-privilege; evidence of non-production data masking; TLS certificate and encryption-at-rest configuration reviewed"),
    ("7.3","Compliance & Security","Are any compliance risks (GDPR / PII / HIPAA / PCI / OSS licences) identified and managed?","Data Protection Impact Assessment (DPIA) or compliance risk register; OSS licence scan report (FOSSA/Black Duck); PII data inventory map if applicable; compliance items in RAID log with owner"),
    ("7.4","Compliance & Security","Are mandatory trainings, security guidelines, and policies adhered to by the team?","Training completion records for all team members (security awareness, GDPR, data handling); signed policy acknowledgement forms; no overdue mandatory training; new joiners trained within 30 days"),
    ("7.5","Compliance & Security","Is evidence for audits (review records, logs, sign-offs) captured and stored properly?","Named folder structure or Confluence space accessible to audit team; review records and sign-offs stored with timestamps; retention period met per policy; audit trail exportable on demand"),
    ("8.1","Financial Health","Is the project on budget (actuals vs planned, effort variance within agreed thresholds)?","Financial tracking report current to last billing period; effort variance within agreed threshold (e.g., +/-10%); EVM or EAC report showing cost performance index (CPI) >= 0.9"),
    ("8.2","Financial Health","Is billing accurate, timely, and aligned with contract / SoW terms?","Invoice copies or billing run report for current month; billing reconciliation against SoW milestones or time-and-material actuals; no disputed invoices older than 30 days without resolution plan"),
    ("8.3","Financial Health","Are margins and utilization aligned with targets and monitored regularly?","Margin report from PSA tool (Salesforce/Replicon) showing billable vs non-billable split; utilization dashboard with actuals vs targets; no resource over/under-allocated beyond agreed thresholds"),
    ("8.4","Financial Health","Are scope / change decisions evaluated for commercial impact before commitment?","Change request log showing commercial impact assessment column populated; approved CRs with account manager sign-off; no uncommercialised scope changes executed in last 30 days"),
    ("8.5","Financial Health","Are there any forecasted financial risks (overruns, discounts, unapproved change orders)?","Financial risk entries in RAID log with owner and mitigation; EAC showing within agreed variance of budget; no unapproved change orders in active execution; overrun escalation evidence if applicable"),
    ("9.1","Customer Success & Growth","What is the latest CSAT / NPS / qualitative feedback from the customer?","Latest CSAT survey results or NPS score from customer (score >= defined threshold, e.g., CSAT >= 4/5); qualitative feedback email or meeting notes from last QBR or monthly governance call"),
    ("9.2","Customer Success & Growth","Is the customer getting measurable value and business outcomes from the engagement?","Benefits realisation report or OKR/KPI tracker updated this quarter; business outcome metrics reported in last status call or steering committee deck; value delivery evidence tied to agreed KPIs"),
    ("9.3","Customer Success & Growth","Are we actively managing key stakeholders and influencing positively at multiple levels?","Stakeholder map with engagement status and last contact date; executive sponsor meeting minutes from last 4 weeks; influence/interest matrix reviewed at last quarterly business review"),
    ("9.4","Customer Success & Growth","Are there identified upsell, cross-sell, or expansion opportunities with clear next steps?","Account growth plan or opportunity register in CRM with pipeline entries; qualified opportunities with next steps, owner, and expected close date; account plan reviewed this quarter"),
    ("9.5","Customer Success & Growth","Are we building references, success stories, or case studies with customer consent?","Customer reference agreement or consent email; draft case study or testimonial in content pipeline with customer approval; NPS promoter score or reference call completed and documented"),
    ("10.1","Continuous Improvement & Knowledge Management","Are regular retrospectives / lessons-learned sessions conducted and actions tracked?","Retrospective records in Jira/Confluence for last 3+ sprints with action items, owners, and status; previous retro actions closed or carried forward with update; improvement trend visible over time"),
    ("10.2","Continuous Improvement & Knowledge Management","Are best practices, templates, and checklists reused and improved across sprints / releases?","Template library in Confluence/SharePoint with version history; evidence of templates updated based on project feedback in last quarter; retrospective actions referencing template improvements"),
    ("10.3","Continuous Improvement & Knowledge Management","Is project knowledge (designs, decisions, FAQs, runbooks) stored in a shared repository and kept current?","Confluence space or SharePoint site with architecture decisions, FAQs, and runbooks; pages updated within last 30 days; all team members have access; search index enabled"),
    ("10.4","Continuous Improvement & Knowledge Management","Are metrics (quality, speed, stability, customer KPIs) used to drive continuous improvement?","Metrics dashboard (Jira/Power BI/Grafana) with quality, velocity, and customer KPIs visible; metrics referenced in retrospectives or steering committee to inform process changes"),
    ("10.5","Continuous Improvement & Knowledge Management","Are transition, handover, and long-term maintainability aspects planned from early stages?","Transition plan covering handover milestones, knowledge transfer schedule, and go-live support model; operational readiness checklist approved by operations or support team"),
]

thin = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
AREA_FILL = PatternFill("solid", fgColor="D6E4F0")
ROW1_FILL = PatternFill("solid", fgColor="EBF3FB")
ROW2_FILL = PatternFill("solid", fgColor="FFFFFF")

def hdr_font():  return Font(name="Calibri", bold=True, color="FFFFFF", size=11)
def area_font(): return Font(name="Calibri", bold=True, color="1F4E79", size=10)
def body_font(): return Font(name="Calibri", size=10)
def wrap_top():  return Alignment(wrap_text=True, vertical="top")
def centered():  return Alignment(wrap_text=True, vertical="center", horizontal="center")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Delivery Check List V 1.0"

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 60
ws.column_dimensions["E"].width = 35
ws.column_dimensions["F"].width = 10

for col, h in enumerate(["SNO","Area","Key Review Question","Expected Evidence","Reviewer Comments","RAG"], 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font(); c.fill = HDR_FILL; c.alignment = centered(); c.border = BORDER
ws.row_dimensions[1].height = 22

for r, (sno, area, question, evidence) in enumerate(DELIVERY_ITEMS, 2):
    fill = ROW1_FILL if r % 2 == 0 else ROW2_FILL
    for col, val in enumerate([sno, area, question, evidence, "", ""], 1):
        c = ws.cell(row=r, column=col, value=val)
        c.border = BORDER
        if col == 1:
            c.font = body_font(); c.alignment = centered(); c.fill = fill
        elif col == 2:
            c.font = area_font(); c.alignment = wrap_top(); c.fill = AREA_FILL
        else:
            c.font = body_font(); c.alignment = wrap_top(); c.fill = fill
    ws.row_dimensions[r].height = 52

ws.freeze_panes = "C2"

out = "c:/projects/nexus-ai/reviewbot/data/templates/reviewbot/v1/standard-delivery.xlsx"
wb.save(out)
print(f"Saved {out}  ({len(DELIVERY_ITEMS)} items)")
