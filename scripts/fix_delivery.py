"""Fix corrupted standard-delivery.xlsx by writing from hardcoded source data."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path

DELIVERY_PATH = Path("data/templates/reviewbot/v1/standard-delivery.xlsx")

HEADERS = [
    "Area", "Category", "Team Category",
    "Question", "Guidance", "Applicability Tags",
    "Evidence", "Weight", "Review?",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ROW_FILL_A = PatternFill("solid", fgColor="EEF2F7")
ROW_FILL_B = PatternFill("solid", fgColor="FFFFFF")
ROW_FONT = Font(name="Calibri", size=10)
ROW_ALIGN = Alignment(vertical="top", wrap_text=True)

# 49 delivery rows: (area, question, evidence)
DELIVERY_ROWS = [
    # Scope, Planning & Governance
    ("Scope, Planning & Governance",
     "Are scope / SoW / baselines clearly defined, signed off, and tracked?",
     "Signed SoW, CR log, scope baseline document with approval trail"),
    ("Scope, Planning & Governance",
     "Are change requests, assumptions, and dependencies captured and actively managed?",
     "CR register with decision dates, assumption log with owners, dependency tracker"),
    ("Scope, Planning & Governance",
     "Is the project plan/schedule maintained and up to date?",
     "Project plan with last-updated date, milestone tracker, critical path"),
    ("Scope, Planning & Governance",
     "Are regular status reports and steering committee meetings conducted?",
     "Status report distribution records, steering committee attendance and minutes"),
    ("Scope, Planning & Governance",
     "Are project decisions, action items and meeting outcomes tracked?",
     "RAID log or action tracker with owner names, due dates, and closure evidence"),
    # Delivery Performance
    ("Delivery Performance",
     "Are project milestones being achieved on schedule?",
     "Burndown chart, milestone tracker, schedule variance report from PM tool"),
    ("Delivery Performance",
     "Is customer acceptance being obtained at agreed checkpoints?",
     "Customer acceptance sign-off emails or documents, review meeting records"),
    ("Delivery Performance",
     "Are blockers and impediments being managed, and is technical debt under control?",
     "Impediment log with resolution dates, technical debt backlog, sprint rework metrics"),
    ("Delivery Performance",
     "Is there a clear release plan with defined scope, date, and stakeholder agreement?",
     "Release plan document with scope, date, dependency list, and stakeholder sign-off"),
    # Requirements Management
    ("Requirements Management",
     "Is the product backlog prioritized, up to date, and with acceptance criteria?",
     "Prioritized backlog with acceptance criteria, version history, approval evidence"),
    ("Requirements Management",
     "Are NFRs (Non-Functional Requirements) defined and approved?",
     "NFR document with quantified SLOs, performance targets, and stakeholder sign-off"),
    ("Requirements Management",
     "Are sprint demos/reviews being conducted with business stakeholders?",
     "Sprint demo recordings or minutes confirming business stakeholder attendance"),
    ("Requirements Management",
     "Is the customer being kept informed through regular reporting?",
     "Customer status reports, distribution list matching agreed stakeholder map"),
    ("Requirements Management",
     "Is customer satisfaction being monitored, and are escalations managed proactively?",
     "CSAT or NPS scores, escalation register with resolution records and owners"),
    # Stakeholder Management
    ("Stakeholder Management",
     "Is there a RAID log that is actively maintained?",
     "RAID log with owners, probability/impact ratings, and last-reviewed dates within 2 weeks"),
    ("Stakeholder Management",
     "Are risk mitigations and contingencies defined for high-probability risks?",
     "Risk mitigation plans and contingency actions documented in RAID log"),
    ("Stakeholder Management",
     "Are issues being resolved within agreed SLAs?",
     "Issue tracker with SLA compliance report and defect age data"),
    ("Stakeholder Management",
     "Is there a formal escalation matrix in place and understood by all?",
     "Escalation matrix document accessible to team and customer, usage evidence"),
    ("Stakeholder Management",
     "Are lessons learned being captured and applied?",
     "Retrospective notes, lessons-learned register, process improvement log"),
    # Risk & Issue Management
    ("Risk & Issue Management",
     "Are resources allocated and available as per the plan?",
     "Resource plan, skills matrix, allocation report vs sprint commitments"),
    ("Risk & Issue Management",
     "Are team members committed and performing at expected capacity?",
     "Jira velocity chart, sprint commitment vs delivery data for last 3 sprints"),
    ("Risk & Issue Management",
     "Are key-person dependencies mitigated with knowledge sharing?",
     "Knowledge transfer documents, backup resource plan, accessible runbooks"),
    ("Risk & Issue Management",
     "Are roles and responsibilities clearly defined and communicated?",
     "RACI matrix with no shared accountabilities, onboarding checklist for new joiners"),
    ("Risk & Issue Management",
     "Is team morale and engagement being monitored?",
     "Team health survey results, retrospective sentiment scores, absence trend data"),
    # People & Resource Management
    ("People & Resource Management",
     "Are quality standards, testing metrics, and code coverage thresholds being met?",
     "CI pipeline test reports, coverage reports, defect density trend chart"),
    ("People & Resource Management",
     "Is UAT being conducted with proper sign-off and defect tracking?",
     "Signed UAT test plan, sign-off document, post-UAT defect leakage report"),
    ("People & Resource Management",
     "Are automated testing frameworks and CI pipelines in place?",
     "CI pipeline configuration, automation coverage report, flaky test tracker"),
    ("People & Resource Management",
     "Are code reviews and static analysis being enforced?",
     "Branch protection settings, peer review records, SonarQube or equivalent scan report"),
    ("People & Resource Management",
     "Are defect trends monitored and addressed?",
     "Defect trend chart, age report, RCA documents for last two critical defects"),
    # Quality Management
    ("Quality Management",
     "Is the project following the agreed quality assurance and governance standards?",
     "SQA checklist completion, audit records, document approval timestamps"),
    ("Quality Management",
     "Are data protection, privacy, and security controls implemented?",
     "Access control matrix, encryption configuration evidence, DPIA or risk register"),
    ("Quality Management",
     "Are regulatory, legal, and compliance requirements identified and managed?",
     "GDPR/PCI compliance register, OSS licence scan report, regulatory risk entries in RAID"),
    ("Quality Management",
     "Have all team members completed mandatory training and certifications?",
     "Training completion records, policy acknowledgement signatures, compliance dashboard"),
    ("Quality Management",
     "Is evidence of compliance and governance controls maintained and accessible?",
     "Audit trail with timestamps, evidence repository with tamper-resistant storage and retention policy"),
    # Compliance & Security
    ("Compliance & Security",
     "Is the project tracking within budget with up-to-date financial reporting?",
     "Financial tracking report, budget variance analysis, EAC vs planned cost"),
    ("Compliance & Security",
     "Is invoicing aligned with contractual milestones and T&M actuals?",
     "Invoice copies matched to SoW milestones, dispute resolution records"),
    ("Compliance & Security",
     "Is resource utilization tracked and billed appropriately?",
     "Utilisation dashboard, billable vs non-billable time reports, over/under allocation flags"),
    ("Compliance & Security",
     "Are scope changes commercially managed through formal change orders?",
     "CR log with commercial impact assessments and account manager sign-off on all CRs"),
    ("Compliance & Security",
     "Are financial risks identified and escalated?",
     "Financial risk register, EAC variance report, escalation evidence if threshold breached"),
    # Financial Management
    ("Financial Management",
     "Is customer satisfaction measured and reported regularly?",
     "CSAT or NPS scores with raw data, QBR slides with customer qualitative feedback"),
    ("Financial Management",
     "Are business outcomes and KPIs being tracked and realized?",
     "Benefits realisation tracker, OKR dashboard, KPI agreement document with customer"),
    ("Financial Management",
     "Are all key stakeholders actively engaged throughout the project?",
     "Stakeholder engagement log, last contact dates, multi-level engagement evidence"),
    ("Financial Management",
     "Are account growth opportunities being identified and pursued?",
     "Account growth plan or CRM opportunity register with named owners and next actions"),
    ("Financial Management",
     "Is the project positioned to generate references, case studies, or upsell opportunities?",
     "Customer reference agreement, case study draft, NPS promoter status evidence"),
    # Customer Value & Satisfaction
    ("Customer Value & Satisfaction",
     "Are retrospectives conducted regularly with actionable outcomes?",
     "Retrospective records with action items, closure evidence, and carry-forward log"),
    ("Customer Value & Satisfaction",
     "Are project processes and templates being improved based on lessons learned?",
     "Template version history, process improvement log with change evidence and feedback source"),
    ("Customer Value & Satisfaction",
     "Is project documentation maintained and accessible to all team members?",
     "Confluence or SharePoint update dates, team access confirmation, document index"),
    ("Customer Value & Satisfaction",
     "Is data-driven decision-making embedded in the delivery process?",
     "Metrics dashboard with quality, velocity, and customer KPIs, data-driven retrospective evidence"),
    ("Customer Value & Satisfaction",
     "Is there a transition/handover plan with operational readiness criteria defined?",
     "Transition plan document with handover milestones, knowledge transfer schedule, ops readiness checklist"),
]

# 49 enrichment entries: (team_category, guidance, applicability_tags, weight)
DELIVERY_ENRICHMENT = [
    # Scope, Planning & Governance
    ("Delivery",
     "Check the project repository for a signed SoW. Ask when scope was last baselined. "
     "Verify the change log shows every accepted CR with a decision date and outcome. "
     "Spot-check two recent changes to confirm they followed the formal CR process.",
     "all", 2.0),
    ("Delivery",
     "Pull the CR log and check for items older than 5 business days without a decision. "
     "Confirm assumptions have owners and are flagged if invalidated. "
     "Verify dependency entries link to the relevant milestone or deliverable.",
     "all", 2.0),
    ("Delivery",
     "Open the project plan in the PM tool. Confirm the last-update date is within 2 weeks. "
     "Check that the critical path is visible and milestones are realistic relative to actuals. "
     "Ask: was the plan shared with stakeholders after the last update?",
     "all", 1.0),
    ("Delivery",
     "Request the last 4 weeks of status reports or steering committee decks. "
     "Check they follow the agreed cadence. Verify action items carry over with owners. "
     "Confirm steering committee attendance includes the required sponsors.",
     "all", 1.0),
    ("Delivery",
     "Open the RAID log or decision register. Check every action has an owner and due date. "
     "Flag any overdue items without an escalation note. "
     "Verify the log was updated within the last 7 days.",
     "all", 1.0),
    # Delivery Performance
    ("Delivery",
     "Review the burndown or velocity chart for the last 3 sprints. "
     "Calculate schedule variance: (planned milestones hit) / (total milestones due). "
     "Ask the PM to walk through any milestone that slipped and what mitigations are in place.",
     "all, agile", 1.0),
    ("Delivery",
     "Check the PR or review history for the last sprint. Confirm no code merged without a peer review. "
     "Review 3 recent customer acceptance emails or comments to verify the customer is actually signing off -- not just rubber-stamping.",
     "all", 2.0),
    ("Delivery",
     "Check the blocker board or impediment log. Verify each blocker has an owner and an expected resolution date. "
     "Review technical-debt backlog size and trend. "
     "Ask how much rework occurred this sprint and whether it is captured in the plan.",
     "all", 1.0),
    ("Delivery",
     "Request the release plan document or Confluence page. "
     "Confirm scope, date, and dependency list are present and agreed. "
     "Ask whether stakeholder sign-off is documented -- verbal is not sufficient.",
     "all", 1.0),
    # Requirements Management
    ("Client",
     "Open the backlog in the PM tool. Spot-check 5 stories in the current sprint: "
     "do all have acceptance criteria? Are they prioritised with a business rationale? "
     "Verify the backlog document is versioned in Git or Confluence with an approval trail.",
     "all", 1.0),
    ("Client",
     "Check that acceptance criteria exist on every story in scope for the current sprint. "
     "Review the NFR document -- are performance, security, and reliability SLOs quantified? "
     "Ask for customer sign-off evidence on the NFR doc.",
     "all", 2.0),
    ("Client",
     "Review the last 3 sprint review recordings, slides, or meeting minutes. "
     "Check that business stakeholders were present (not just the project team). "
     "Verify that demo feedback was captured and actioned in the following sprint.",
     "all, agile", 1.0),
    ("Client",
     "Review the last 4 status reports sent to the customer. "
     "Check they match the agreed cadence and format. "
     "Confirm the distribution list includes all named stakeholders from the stakeholder map.",
     "all", 1.0),
    ("Client",
     "Ask directly: have there been any escalation emails or formal complaints from the customer this quarter? "
     "Check the risk log for any customer-relationship risk entries. "
     "If escalations exist, verify resolution actions have an owner and a target date.",
     "all", 2.0),
    # Stakeholder Management
    ("Governance",
     "Open the RAID log. Verify every entry has an owner, probability, impact, and last-reviewed date. "
     "Flag any entry not reviewed in the last 2 weeks as stale. "
     "Ask whether the RAID log is reviewed in the weekly project call or only updated ad hoc.",
     "all", 2.0),
    ("Governance",
     "Filter the RAID log to Red and Amber risks. Check that every high-risk item has a mitigation AND a contingency. "
     "Ask the project manager to describe one contingency plan verbally -- can they do so without looking it up?",
     "all", 2.0),
    ("Governance",
     "Pull the issue tracker. Calculate the average age of open P1 and P2 issues. "
     "Compare against the agreed SLA. Flag any issue past SLA without an escalation record. "
     "Check the issue closure rate trend over the last 3 sprints.",
     "all", 2.0),
    ("Governance",
     "Ask the team to describe the escalation matrix without prompting. "
     "Check that the matrix is documented and accessible to both internal team and customer. "
     "Verify at least one real escalation in the last quarter was handled through the documented process.",
     "all", 1.0),
    ("Governance",
     "Review retrospective action items for recurring themes across the last 3 sprints. "
     "Check whether lessons-learned entries have fed back into the project plan or process. "
     "Ask: what one recurring risk has improved as a result of a process change?",
     "all", 1.0),
    # Risk & Issue Management
    ("Resourcing",
     "Review the resource plan: does the skills matrix map individuals to the deliverables they own? "
     "Check allocation percentages against actual sprint commitments. "
     "Ask whether any critical skill gap is currently open and what the mitigation plan is.",
     "all", 2.0),
    ("Resourcing",
     "Check Jira velocity over the last 3 sprints -- is it stable or declining? "
     "Ask the project manager whether any planned resource has become unavailable. "
     "Verify RAID log has a resource-risk entry with mitigation if availability is below plan.",
     "all", 1.0),
    ("Resourcing",
     "Identify the 3 most critical roles. For each, confirm a named backup person exists in the resource plan. "
     "Check knowledge-transfer documentation is accessible for each critical role -- not just in one person's head.",
     "all", 2.0),
    ("Resourcing",
     "For each team member who joined in the last 3 months, verify an onboarding checklist was completed. "
     "Check the RACI for clarity: no more than one Accountable per deliverable, Responsible clearly named. "
     "Ask a recent joiner whether they felt equipped to contribute in their first two weeks.",
     "all", 1.0),
    ("Resourcing",
     "Review retrospective sentiment scores or team health survey results from the last 3 sprints. "
     "Check absence and leave patterns for any signal of disengagement. "
     "Ask the delivery manager privately whether there are any unresolved interpersonal issues.",
     "all", 1.0),
    # People & Resource Management
    ("QA",
     "Open the CI pipeline. Check test pass rate and code coverage against the agreed threshold. "
     "Review the defect density metric -- is it trending down? "
     "Ask: what is the lowest coverage module, and is there a plan to improve it?",
     "all", 2.0),
    ("QA",
     "Check for a signed UAT test plan and a UAT sign-off document or email. "
     "Calculate post-UAT defect leakage: how many P1/P2 defects reached production after UAT sign-off? "
     "Any leakage above the agreed threshold should be treated as a critical quality gap.",
     "all", 2.0),
    ("QA",
     "Review the CI pipeline configuration. Verify automated tests run on every PR. "
     "Check the automation coverage report -- is the regression suite meaningfully automated or mostly manual? "
     "Ask whether flaky tests are tracked and addressed.",
     "all", 1.0),
    ("QA",
     "Check branch protection settings: is at least one peer review required before merge? "
     "Review the SonarQube or equivalent scan report -- any unresolved critical issues? "
     "Ask for the last performance and security test results and compare to NFR thresholds.",
     "all", 2.0),
    ("QA",
     "Pull the defect trend chart. Is the open-defect count stable or growing? "
     "Check defect age -- are any defects older than 2 sprints without resolution? "
     "Ask for the RCA on the last two critical defects -- were root causes addressed or just symptoms?",
     "all", 1.0),
    # Quality Management
    ("Governance",
     "Check the SQA checklist completion status for the current delivery phase. "
     "Verify document approval records include timestamps and named approvers. "
     "Ask when the last internal audit occurred and whether findings are all closed.",
     "all", 2.0),
    ("Governance",
     "Review the access control matrix: is least-privilege enforced? "
     "Ask for evidence that non-production environments use masked or synthetic data, not production copies. "
     "Verify TLS is configured and encryption-at-rest is enabled for all data stores.",
     "all, regulated", 3.0),
    ("Governance",
     "Check whether a DPIA or compliance risk register exists and is current. "
     "Run a quick scan for GDPR/PII/PCI items in the RAID log -- are all owned and mitigated? "
     "Ask for the OSS licence scan report and verify no GPL or AGPL licences in a proprietary context.",
     "all, regulated", 3.0),
    ("Governance",
     "Pull training completion records. Has every team member completed mandatory security awareness training? "
     "Check for any overdue training -- especially for team members who joined in the last 90 days. "
     "Verify policy acknowledgement signatures are on file.",
     "all", 2.0),
    ("Governance",
     "Ask where audit evidence is stored. Can it be retrieved in under 5 minutes? "
     "Check that review records include timestamps and are tamper-resistant. "
     "Confirm the retention period for all evidence meets policy requirements.",
     "all", 2.0),
    # Compliance & Security
    ("Delivery",
     "Request the latest financial tracking report. Is the effort variance within +/-10% of plan? "
     "Check CPI or EAC -- is the project forecast to complete within budget? "
     "Ask: what is driving any variance, and is it a systemic issue or a one-off?",
     "all", 2.0),
    ("Delivery",
     "Review the last two invoices: do they match the SoW milestones or T&M actuals? "
     "Check for any disputed invoices older than 30 days without a documented resolution plan. "
     "Ask the account manager when the next billing milestone is due and whether it is at risk.",
     "all", 1.0),
    ("Delivery",
     "Pull the utilization dashboard. Is billable utilization at or above target? "
     "Check non-billable time -- is it within agreed allowances? "
     "Flag any resource over-allocated (> 100%) or significantly under-allocated for more than 2 weeks.",
     "all", 1.0),
    ("Delivery",
     "Check the CR log for commercial impact assessment on each change. "
     "Ask whether any scope additions were delivered without a signed CR. "
     "Verify account manager sign-off is present on all approved CRs.",
     "all", 2.0),
    ("Delivery",
     "Review financial risks in the RAID log. Is EAC within the agreed variance band? "
     "Ask directly: are there any unapproved change orders in active delivery? "
     "Check whether overrun risk has been escalated to the sponsor if EAC exceeds threshold.",
     "all", 2.0),
    # Financial Management
    ("Client",
     "Ask for the most recent CSAT score or NPS result with the raw data, not just the headline number. "
     "Review qualitative feedback from the last QBR. "
     "Ask: is there any feedback that has not yet been actioned?",
     "all", 1.0),
    ("Client",
     "Request the benefits realisation or OKR tracker. Are business outcomes measured and improving? "
     "Ask the customer (or the engagement lead) to describe one concrete business benefit delivered this quarter. "
     "Verify KPIs are agreed, not just tracked internally.",
     "all", 1.0),
    ("Client",
     "Review the stakeholder map: when was each key stakeholder last engaged? "
     "Ask whether the engagement includes multiple levels (operational, management, executive). "
     "Check that no key stakeholder has been inactive for more than 30 days.",
     "all", 1.0),
    ("Client",
     "Open the account growth plan or CRM opportunity register. "
     "Check that any identified opportunities have a named owner and a next action with a date. "
     "Ask the engagement lead: is the customer aware we are interested in expanding the partnership?",
     "all", 1.0),
    ("Client",
     "Ask whether a customer reference agreement or consent is on file. "
     "Check whether a draft case study or testimonial exists and has been shared with the customer. "
     "Verify NPS promoter status -- would this customer actively recommend us?",
     "all", 1.0),
    # Customer Value & Satisfaction
    ("Delivery",
     "Review retrospective records for the last 3 sprints. "
     "Check that every action item from the previous retro has been closed or carried forward with an update. "
     "Ask: what one practice changed as a direct result of a retrospective?",
     "all", 1.0),
    ("Delivery",
     "Check the template library or Confluence space. Are templates versioned and updated based on project feedback? "
     "Ask the project manager to show you a template that was improved in the last quarter -- and what prompted the change.",
     "all", 1.0),
    ("Delivery",
     "Open the project Confluence space or SharePoint site. "
     "Check when pages were last updated -- any page older than 30 days should be flagged for review. "
     "Verify that all current team members can find and access key documents without escalation.",
     "all", 1.0),
    ("Delivery",
     "Ask the team to show you the metrics dashboard used in retrospectives or steering committee. "
     "Check that at least one process change can be traced back to a data-driven decision. "
     "Verify quality, velocity, and customer KPIs are all visible in one place.",
     "all", 1.0),
    ("Delivery",
     "Ask for the transition plan. Does it cover handover milestones, knowledge transfer schedule, and go-live support model? "
     "Check that the operational readiness checklist has been reviewed by the operations or support team. "
     "Verify transition planning started before the final delivery phase -- not at the last sprint.",
     "all", 2.0),
]

assert len(DELIVERY_ROWS) == 49, f"Expected 49 rows, got {len(DELIVERY_ROWS)}"
assert len(DELIVERY_ENRICHMENT) == 49, f"Expected 49 enrichment, got {len(DELIVERY_ENRICHMENT)}"

# Build workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Master Template Items"
ws.freeze_panes = "A2"
ws.append(HEADERS)

for (area, question, evidence), (team, guidance, tags, weight) in zip(DELIVERY_ROWS, DELIVERY_ENRICHMENT):
    ws.append([area, "Delivery", team, question, guidance, tags, evidence, weight, "Yes"])

# Style header
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN

# Style data rows
for i, row in enumerate(ws.iter_rows(min_row=2), 1):
    fill = ROW_FILL_A if i % 2 == 0 else ROW_FILL_B
    for cell in row:
        cell.fill = fill
        cell.font = ROW_FONT
        cell.alignment = ROW_ALIGN

# Column widths
widths = {"A": 26, "B": 12, "C": 16, "D": 42, "E": 52, "F": 22, "G": 52, "H": 8, "I": 8}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 28

wb.save(DELIVERY_PATH)
print(f"Written {ws.max_row - 1} data rows to {DELIVERY_PATH}")

# Verify
wb2 = openpyxl.load_workbook(DELIVERY_PATH)
ws2 = wb2.active
r2 = list(ws2.iter_rows(min_row=2, max_row=2, values_only=True))[0]
print(f"Row 2 check: area={r2[0]}, cat={r2[1]}, team={r2[2]}, q={r2[3][:60]}")
print("DONE")
