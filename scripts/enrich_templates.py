"""
One-time script to enrich source Excel templates with full column set.

Run from reviewbot directory:
    python scripts/enrich_templates.py

Adds columns: Category, Team Category, Guidance, Applicability Tags, Weight, Review?
Renames: Key Review Question → Question, Expected Evidence → Evidence
Drops: SNO/#, Reviewer Comments, RAG (runtime review fields)
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

DELIVERY_PATH = Path("data/templates/reviewbot/v1/standard-delivery.xlsx")
TECHNICAL_PATH = Path("data/templates/reviewbot/v1/standard-technical.xlsx")

HEADERS = [
    "Area", "Category", "Team Category",
    "Question", "Guidance", "Applicability Tags",
    "Evidence", "Weight", "Review?",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_FILL_A = PatternFill("solid", fgColor="EEF2F7")
ROW_FILL_B = PatternFill("solid", fgColor="FFFFFF")
ROW_FONT = Font(name="Calibri", size=10)
ROW_ALIGN = Alignment(vertical="top", wrap_text=True)

# Each tuple: (team_category, guidance, applicability_tags, weight)
# Row order matches existing source files.

DELIVERY_ENRICHMENT = [
    # ── Scope, Planning & Governance ─────────────────────────────────────────
    ("Delivery",
     "Check the project repository for a signed SoW. Ask when scope was last baselined. "
     "Verify the change log shows every accepted CR with a decision date and outcome. "
     "Spot-check two recent changes to confirm they followed the formal CR process.",
     "all", 2.0),
    ("Delivery",
     "Pull the CR log and check for items older than 5 business days without a decision. "
     "Confirm assumptions have owners and are flagged if invalidated. "
     "Verify dependency entries link to the relevant milestone or deliverable.",
     "all", 2.0),
    ("Delivery",
     "Open the project plan in the PM tool. Confirm the last-update date is within 2 weeks. "
     "Check that the critical path is visible and milestones are realistic relative to actuals. "
     "Ask: was the plan shared with stakeholders after the last update?",
     "all", 1.0),
    ("Delivery",
     "Request the last 4 weeks of status reports or steering committee decks. "
     "Check they follow the agreed cadence. Verify action items carry over with owners. "
     "Confirm steering committee attendance includes the required sponsors.",
     "all", 1.0),
    ("Delivery",
     "Open the RAID log or decision register. Check every action has an owner and due date. "
     "Flag any overdue items without an escalation note. "
     "Verify the log was updated within the last 7 days.",
     "all", 1.0),
    # ── Delivery Health ───────────────────────────────────────────────────────
    ("Delivery",
     "Review the burndown or velocity chart for the last 3 sprints. "
     "Calculate schedule variance: (planned milestones hit) / (total milestones due). "
     "Ask the PM to walk through any milestone that slipped and what mitigations are in place.",
     "all, agile", 1.0),
    ("Delivery",
     "Check the PR or review history for the last sprint. Confirm no code merged without a peer review. "
     "Review 3 recent customer acceptance emails or comments to verify the customer is actually signing off — not just rubber-stamping.",
     "all", 2.0),
    ("Delivery",
     "Check the blocker board or impediment log. Verify each blocker has an owner and an expected resolution date. "
     "Review technical-debt backlog size and trend. "
     "Ask how much rework occurred this sprint and whether it is captured in the plan.",
     "all", 1.0),
    ("Delivery",
     "Request the release plan document or Confluence page. "
     "Confirm scope, date, and dependency list are present and agreed. "
     "Ask whether stakeholder sign-off is documented — verbal is not sufficient.",
     "all", 1.0),
    # ── Requirements & Customer Alignment ─────────────────────────────────────
    ("Client",
     "Open the backlog in the PM tool. Spot-check 5 stories in the current sprint: "
     "do all have acceptance criteria? Are they prioritised with a business rationale? "
     "Verify the backlog document is versioned in Git or Confluence with an approval trail.",
     "all", 1.0),
    ("Client",
     "Check that acceptance criteria exist on every story in scope for the current sprint. "
     "Review the NFR document — are performance, security, and reliability SLOs quantified? "
     "Ask for customer sign-off evidence on the NFR doc.",
     "all", 2.0),
    ("Client",
     "Review the last 3 sprint review recordings, slides, or meeting minutes. "
     "Check that business stakeholders were present (not just the project team). "
     "Verify that demo feedback was captured and actioned in the following sprint.",
     "all, agile", 1.0),
    ("Client",
     "Review the last 4 status reports sent to the customer. "
     "Check they match the agreed cadence and format. "
     "Confirm the distribution list includes all named stakeholders from the stakeholder map.",
     "all", 1.0),
    ("Client",
     "Ask directly: have there been any escalation emails or formal complaints from the customer this quarter? "
     "Check the risk log for any customer-relationship risk entries. "
     "If escalations exist, verify resolution actions have an owner and a target date.",
     "all", 2.0),
    # ── Risks, Issues & Escalations ───────────────────────────────────────────
    ("Governance",
     "Open the RAID log. Verify every entry has an owner, probability, impact, and last-reviewed date. "
     "Flag any entry not reviewed in the last 2 weeks as stale. "
     "Ask whether the RAID log is reviewed in the weekly project call or only updated ad hoc.",
     "all", 2.0),
    ("Governance",
     "Filter the RAID log to Red and Amber risks. Check that every high-risk item has a mitigation AND a contingency. "
     "Ask the project manager to describe one contingency plan verbally — can they do so without looking it up?",
     "all", 2.0),
    ("Governance",
     "Pull the issue tracker. Calculate the average age of open P1 and P2 issues. "
     "Compare against the agreed SLA. Flag any issue past SLA without an escalation record. "
     "Check the issue closure rate trend over the last 3 sprints.",
     "all", 2.0),
    ("Governance",
     "Ask the team to describe the escalation matrix without prompting. "
     "Check that the matrix is documented and accessible to both internal team and customer. "
     "Verify at least one real escalation in the last quarter was handled through the documented process.",
     "all", 1.0),
    ("Governance",
     "Review retrospective action items for recurring themes across the last 3 sprints. "
     "Check whether lessons-learned entries have fed back into the project plan or process. "
     "Ask: what one recurring risk has improved as a result of a process change?",
     "all", 1.0),
    # ── Resource & Team Management ────────────────────────────────────────────
    ("Resourcing",
     "Review the resource plan: does the skills matrix map individuals to the deliverables they own? "
     "Check allocation percentages against actual sprint commitments. "
     "Ask whether any critical skill gap is currently open and what the mitigation plan is.",
     "all", 2.0),
    ("Resourcing",
     "Check Jira velocity over the last 3 sprints — is it stable or declining? "
     "Ask the project manager whether any planned resource has become unavailable. "
     "Verify RAID log has a resource-risk entry with mitigation if availability is below plan.",
     "all", 1.0),
    ("Resourcing",
     "Identify the 3 most critical roles. For each, confirm a named backup person exists in the resource plan. "
     "Check knowledge-transfer documentation is accessible for each critical role — not just in one person's head.",
     "all", 2.0),
    ("Resourcing",
     "For each team member who joined in the last 3 months, verify an onboarding checklist was completed. "
     "Check the RACI for clarity: no more than one Accountable per deliverable, Responsible clearly named. "
     "Ask a recent joiner whether they felt equipped to contribute in their first two weeks.",
     "all", 1.0),
    ("Resourcing",
     "Review retrospective sentiment scores or team health survey results from the last 3 sprints. "
     "Check absence and leave patterns for any signal of disengagement. "
     "Ask the delivery manager privately whether there are any unresolved interpersonal issues.",
     "all", 1.0),
    # ── Quality & Testing ─────────────────────────────────────────────────────
    ("QA",
     "Open the CI pipeline. Check test pass rate and code coverage against the agreed threshold. "
     "Review the defect density metric — is it trending down? "
     "Ask: what is the lowest coverage module, and is there a plan to improve it?",
     "all", 2.0),
    ("QA",
     "Check for a signed UAT test plan and a UAT sign-off document or email. "
     "Calculate post-UAT defect leakage: how many P1/P2 defects reached production after UAT sign-off? "
     "Any leakage above the agreed threshold should be treated as a critical quality gap.",
     "all", 2.0),
    ("QA",
     "Review the CI pipeline configuration. Verify automated tests run on every PR. "
     "Check the automation coverage report — is the regression suite meaningfully automated or mostly manual? "
     "Ask whether flaky tests are tracked and addressed.",
     "all", 1.0),
    ("QA",
     "Check branch protection settings: is at least one peer review required before merge? "
     "Review the SonarQube or equivalent scan report — any unresolved critical issues? "
     "Ask for the last performance and security test results and compare to NFR thresholds.",
     "all", 2.0),
    ("QA",
     "Pull the defect trend chart. Is the open-defect count stable or growing? "
     "Check defect age — are any defects older than 2 sprints without resolution? "
     "Ask for the RCA on the last two critical defects — were root causes addressed or just symptoms?",
     "all", 1.0),
    # ── Compliance & Security ─────────────────────────────────────────────────
    ("Governance",
     "Check the SQA checklist completion status for the current delivery phase. "
     "Verify document approval records include timestamps and named approvers. "
     "Ask when the last internal audit occurred and whether findings are all closed.",
     "all", 2.0),
    ("Governance",
     "Review the access control matrix: is least-privilege enforced? "
     "Ask for evidence that non-production environments use masked or synthetic data, not production copies. "
     "Verify TLS is configured and encryption-at-rest is enabled for all data stores.",
     "all, regulated", 3.0),
    ("Governance",
     "Check whether a DPIA or compliance risk register exists and is current. "
     "Run a quick scan for GDPR/PII/PCI items in the RAID log — are all owned and mitigated? "
     "Ask for the OSS licence scan report and verify no GPL or AGPL licences in a proprietary context.",
     "all, regulated", 3.0),
    ("Governance",
     "Pull training completion records. Has every team member completed mandatory security awareness training? "
     "Check for any overdue training — especially for team members who joined in the last 90 days. "
     "Verify policy acknowledgement signatures are on file.",
     "all", 2.0),
    ("Governance",
     "Ask where audit evidence is stored. Can it be retrieved in under 5 minutes? "
     "Check that review records include timestamps and are tamper-resistant. "
     "Confirm the retention period for all evidence meets policy requirements.",
     "all", 2.0),
    # ── Financial Health ──────────────────────────────────────────────────────
    ("Delivery",
     "Request the latest financial tracking report. Is the effort variance within +/-10% of plan? "
     "Check CPI or EAC — is the project forecast to complete within budget? "
     "Ask: what is driving any variance, and is it a systemic issue or a one-off?",
     "all", 2.0),
    ("Delivery",
     "Review the last two invoices: do they match the SoW milestones or T&M actuals? "
     "Check for any disputed invoices older than 30 days without a documented resolution plan. "
     "Ask the account manager when the next billing milestone is due and whether it is at risk.",
     "all", 1.0),
    ("Delivery",
     "Pull the utilization dashboard. Is billable utilization at or above target? "
     "Check non-billable time — is it within agreed allowances? "
     "Flag any resource over-allocated (> 100%) or significantly under-allocated for more than 2 weeks.",
     "all", 1.0),
    ("Delivery",
     "Check the CR log for commercial impact assessment on each change. "
     "Ask whether any scope additions were delivered without a signed CR. "
     "Verify account manager sign-off is present on all approved CRs.",
     "all", 2.0),
    ("Delivery",
     "Review financial risks in the RAID log. Is EAC within the agreed variance band? "
     "Ask directly: are there any unapproved change orders in active delivery? "
     "Check whether overrun risk has been escalated to the sponsor if EAC exceeds threshold.",
     "all", 2.0),
    # ── Customer Success & Growth ─────────────────────────────────────────────
    ("Client",
     "Ask for the most recent CSAT score or NPS result with the raw data, not just the headline number. "
     "Review qualitative feedback from the last QBR. "
     "Ask: is there any feedback that has not yet been actioned?",
     "all", 1.0),
    ("Client",
     "Request the benefits realisation or OKR tracker. Are business outcomes measured and improving? "
     "Ask the customer (or the engagement lead) to describe one concrete business benefit delivered this quarter. "
     "Verify KPIs are agreed, not just tracked internally.",
     "all", 1.0),
    ("Client",
     "Review the stakeholder map: when was each key stakeholder last engaged? "
     "Ask whether the engagement includes multiple levels (operational, management, executive). "
     "Check that no key stakeholder has been inactive for more than 30 days.",
     "all", 1.0),
    ("Client",
     "Open the account growth plan or CRM opportunity register. "
     "Check that any identified opportunities have a named owner and a next action with a date. "
     "Ask the engagement lead: is the customer aware we are interested in expanding the partnership?",
     "all", 1.0),
    ("Client",
     "Ask whether a customer reference agreement or consent is on file. "
     "Check whether a draft case study or testimonial exists and has been shared with the customer. "
     "Verify NPS promoter status — would this customer actively recommend us?",
     "all", 1.0),
    # ── Continuous Improvement & Knowledge Management ─────────────────────────
    ("Delivery",
     "Review retrospective records for the last 3 sprints. "
     "Check that every action item from the previous retro has been closed or carried forward with an update. "
     "Ask: what one practice changed as a direct result of a retrospective?",
     "all", 1.0),
    ("Delivery",
     "Check the template library or Confluence space. Are templates versioned and updated based on project feedback? "
     "Ask the project manager to show you a template that was improved in the last quarter — and what prompted the change.",
     "all", 1.0),
    ("Delivery",
     "Open the project Confluence space or SharePoint site. "
     "Check when pages were last updated — any page older than 30 days should be flagged for review. "
     "Verify that all current team members can find and access key documents without escalation.",
     "all", 1.0),
    ("Delivery",
     "Ask the team to show you the metrics dashboard used in retrospectives or steering committee. "
     "Check that at least one process change can be traced back to a data-driven decision. "
     "Verify quality, velocity, and customer KPIs are all visible in one place.",
     "all", 1.0),
    ("Delivery",
     "Ask for the transition plan. Does it cover handover milestones, knowledge transfer schedule, and go-live support model? "
     "Check that the operational readiness checklist has been reviewed by the operations or support team. "
     "Verify transition planning started before the final delivery phase — not at the last sprint.",
     "all", 2.0),
]

TECHNICAL_ENRICHMENT = [
    # ── Architecture & Design ─────────────────────────────────────────────────
    ("Development",
     "Ask who owns architecture decisions on this project. Is that person available and engaged? "
     "Check the RACI for architecture: is accountability clearly named, not shared? "
     "Verify the named architect has contributed to the design artefacts within the last sprint.",
     "all", 1.0),
    ("Development",
     "Ask when the architecture diagrams were last updated. Do they reflect the actual deployed state? "
     "Ask a developer to walk you through the C4 diagrams without prompting — can they? "
     "Verify diagram source files are in Git, not only as image attachments.",
     "all", 2.0),
    ("Development",
     "Request HLD and LLD for the two most complex modules. "
     "Check commit dates on the docs — are they within 3 months? "
     "Spot-check: does the LLD reference the actual classes or components in the codebase?",
     "all", 2.0),
    ("Development",
     "Review ADRs. Is there at least one per major technology or design choice? "
     "Check that each ADR covers context, decision, and consequences — not just a headline. "
     "Ask: which of these decisions is most at risk of reversal and why?",
     "all", 1.0),
    ("Development",
     "Run a static analysis or review the SonarQube dashboard for coupling and complexity metrics. "
     "Check that the architecture diagram shows clear service boundaries with no undocumented cross-cutting dependencies. "
     "Ask: can a new feature be added to one module without changing another?",
     "all", 1.0),
    ("Development",
     "Ask the architect to name two design patterns used in the codebase and point to where they are applied. "
     "Check whether the reference architecture is aligned to enterprise standards or client guidelines. "
     "Verify that patterns are consistently applied, not just present in one module.",
     "all", 1.0),
    ("Development",
     "Review the folder structure and layer diagram. "
     "Can you trace a feature from UI through API, business logic, and data layers without crossing unexpected boundaries? "
     "Use a dependency analysis tool to verify no cross-layer imports exist.",
     "all", 1.0),
    ("Development",
     "Request the integration architecture diagram. Check that all inbound and outbound APIs, event streams, and ETL flows are represented. "
     "Verify API contracts or AsyncAPI specs exist for each integration and are committed to the repo.",
     "all", 2.0),
    ("Development",
     "Check the architecture doc for a cross-cutting concerns section. "
     "Verify that logging, security, caching, and observability are addressed by shared libraries or middleware — not implemented ad hoc per service.",
     "all", 2.0),
    ("Development",
     "Ask for the NFR document. Are latency, availability, and reliability SLOs quantified with numbers, not qualitative descriptions? "
     "Trace two NFRs to a specific architecture or design decision — does the traceability exist?",
     "all", 2.0),
    ("Development",
     "Check the dependency version list. Are any libraries end-of-life or more than 2 major versions behind? "
     "Verify the tech stack against the enterprise architecture approved list, if one exists. "
     "Flag any component with a known CVE that has not been scheduled for upgrade.",
     "all", 2.0),
    ("Development",
     "Open the technical debt register. Is every deprecated or legacy component listed with a migration plan? "
     "Check that no deprecated component is being newly extended without an approved exception. "
     "Ask: when is the next planned deprecation milestone?",
     "all", 1.0),
    # ── Technical Documentation ───────────────────────────────────────────────
    ("Development",
     "Navigate to /docs/ in the repository. Is the architecture documentation committed and versioned in Git? "
     "Check modification dates on Confluence pages — any page stale beyond 3 months should be flagged. "
     "Ask: when was documentation last verified against actual system behaviour?",
     "all", 1.0),
    ("Development",
     "Locate the OpenAPI/Swagger spec and Postman collection in the repo. "
     "Spot-check 3 endpoints: do the spec and actual API responses match? "
     "For event-driven systems, ask for the AsyncAPI spec — is it in the repo?",
     "all", 2.0),
    ("Development",
     "Ask a new team member (or recent joiner) to follow the README setup steps from scratch on a clean machine — or ask if they have done so. "
     "Check the deployment runbook: does it cover all environments? "
     "Verify config parameters are documented with defaults and whether they are required or optional.",
     "all", 1.0),
    ("Development",
     "Ask to see the developer onboarding guide. Does it cover branching strategy, coding standards, PR process, and local setup? "
     "Verify it has been validated by the most recent joiner. "
     "Check it is in the repo root or /docs/ — not hidden in a personal folder.",
     "all", 1.0),
    ("Development",
     "Ask the operations team (not the development team) to locate the runbook for their most common task. Can they find it without developer help? "
     "Check that user guides and admin runbooks exist and are kept current. "
     "Verify documentation is accessible without requiring repository access.",
     "all", 1.0),
    ("Development",
     "Check that ER diagrams are committed in the repo and match the actual migration scripts. "
     "Verify integration contracts (OpenAPI/AsyncAPI) are up to date — compare a recent API change to its corresponding contract update commit date.",
     "all", 2.0),
    ("Development",
     "Open KNOWN_ISSUES.md or the Confluence technical debt page. "
     "Is every known issue listed with an owner and a target date? "
     "Ask: are there any known production issues not on this list? Flag any unacknowledged defects.",
     "all", 2.0),
    # ── Data & Storage Design ────────────────────────────────────────────────
    ("Development",
     "Review the ER diagram against the actual schema migration files — do they match? "
     "Check for generic table names (e.g., 'data', 'records') that suggest a poor domain model. "
     "Ask: was the schema designed with domain-driven design principles or built ad hoc?",
     "all", 2.0),
    ("Development",
     "Pull EXPLAIN ANALYZE output for the top 5 most frequent queries. "
     "Are there full table scans on tables with more than 100k rows? "
     "Ask: how were indexes chosen, and when were index usage statistics last reviewed?",
     "all", 2.0),
    ("Development",
     "Ask for the query performance analysis report. "
     "Check that slow queries are being identified (slow query log enabled) and tracked in Jira. "
     "Verify that optimized queries were retested and improvement is measured, not assumed.",
     "all", 1.0),
    ("Development",
     "Request the backup policy document. Verify RPO and RTO targets are defined. "
     "Ask: when was the last restore drill performed, and what was the actual recovery time? "
     "Check that the restore procedure is documented, not just the backup schedule.",
     "all", 3.0),
    ("Development",
     "Check the data retention policy. Are retention periods defined per data class? "
     "Verify an automated archival or purging job exists and has been tested in a non-production environment. "
     "Ask: what happens to archived data — is it retrievable if needed for legal hold?",
     "all, regulated", 2.0),
    ("Development",
     "Review the DB access control matrix: does every role follow least privilege? "
     "Check encryption-at-rest is enabled in the cloud console or IaC. "
     "Verify network firewall rules restrict the DB port to the application tier only — no direct public access.",
     "all", 3.0),
    ("Development",
     "Check the IaC for HA configuration (multi-AZ, read replica, or cluster). "
     "Ask for evidence of a failover test: was the switchover time within the agreed RTO? "
     "Verify that a monitoring alert is configured on replica lag.",
     "all, cloud", 2.0),
    ("Development",
     "Request the data migration plan document. Does it include a source-to-target mapping and a reconciliation approach? "
     "Ask: has the migration been rehearsed in a staging environment, and was the data reconciled after? "
     "Verify the rollback procedure is documented and has been tested.",
     "all", 2.0),
    ("Development",
     "Review schema migration scripts for NOT NULL, FK, CHECK, and UNIQUE constraints. "
     "Check the data validation layer in application code — is validation applied before DB writes? "
     "Ask: what happens when a data quality check fails — is it logged, surfaced, or silently ignored?",
     "all", 2.0),
    # ── Security Architecture ─────────────────────────────────────────────────
    ("DevOps",
     "Verify TLS 1.2+ is enforced on all endpoints — run SSL Labs against the primary domain. "
     "Check DB encryption-at-rest in the cloud console or IaC. "
     "Scan the codebase with git-secrets or TruffleHog for any plaintext secrets in config files or Git history.",
     "all, regulated", 3.0),
    ("DevOps",
     "Review the authentication flow diagram in the architecture doc. "
     "Check that SSO/OIDC/OAuth2 is integrated — not a home-grown auth system. "
     "Verify MFA is enabled for privileged accounts: ask for a screenshot from the IAM console, not just confirmation.",
     "all", 3.0),
    ("DevOps",
     "Request the role definition document with the permissions matrix. "
     "Spot-check 3 users in the production system: do their actual permissions match the matrix? "
     "Ask for access review records — have permissions been reviewed in the last 3 months?",
     "all", 3.0),
    ("DevOps",
     "Ask: where are secrets stored in this project? If the answer is anything other than a vault, flag it immediately. "
     "Verify git-secrets or a pre-commit hook is active. "
     "Run a SAST scan and check for any hardcoded credential findings.",
     "all", 3.0),
    ("DevOps",
     "Check the codebase for an input validation library (Pydantic, Joi, class-validator). "
     "Run securityheaders.com against the application URL and verify grade A. "
     "Ask for the OWASP Dependency Check report — any critical findings?",
     "all", 2.0),
    ("DevOps",
     "Request the SAST report (SonarQube/Checkmarx/Semgrep). Are there any unresolved critical or high findings? "
     "If a DAST scan (OWASP ZAP) has been run against staging, review the results. "
     "Ask: when was the last penetration test, and are all findings from it closed?",
     "all", 3.0),
    ("DevOps",
     "Request the SCA/SBOM report (Snyk/Dependabot). Are there any unresolved critical CVEs in dependencies? "
     "Check the CI pipeline for a container image scan stage (Trivy/Grype). "
     "Verify an automated update policy (Dependabot or Renovate) is active.",
     "all", 3.0),
    ("DevOps",
     "Check that TLS certificates are managed in the load balancer or ingress IaC, not manually. "
     "Verify HTTP-to-HTTPS redirect is configured. "
     "Ask: is there a certificate expiry monitoring alert with at least 30 days' lead time?",
     "all, cloud", 2.0),
    ("DevOps",
     "Review a sample of recent log entries from the logging platform. "
     "Check that auth events, privilege changes, and data access are logged. "
     "Verify that no PII, secrets, or tokens appear in log lines — sample at least 20 records.",
     "all", 2.0),
    ("DevOps",
     "Check the security testing plan in the project schedule. "
     "Verify SAST/DAST results are available in the CI pipeline with a gating policy. "
     "Ask: are security findings tracked in Jira with an owner and target date, or just acknowledged and ignored?",
     "all", 2.0),
    # ── Code Quality & Standards ─────────────────────────────────────────────
    ("Development",
     "Check the repo for CONTRIBUTING.md and linter config files. "
     "Verify the CI pipeline fails on critical linter violations. "
     "Spot-check the last 5 merged PRs: were linter violations present and addressed?",
     "all", 1.0),
    ("Development",
     "Check branch protection rules in the repo settings. Is at least one reviewer required before merge? "
     "Review the last 20 merged PRs for evidence of meaningful review comments — not just approvals. "
     "Ask: who is responsible for enforcing review quality?",
     "all", 2.0),
    ("Development",
     "Run git-secrets or TruffleHog across the full Git history. "
     "Check the .env.example file — does it show all required variables without real values? "
     "Verify config is injected via environment variables or a vault at runtime, not committed as flat files.",
     "all", 3.0),
    ("Development",
     "Ask to see the central error handler or middleware. Is it applied consistently across all services? "
     "Check the error response schema in the OpenAPI spec — are error codes meaningful? "
     "Review the application logs for the last 7 days: any unhandled exceptions or silent failures?",
     "all", 2.0),
    ("Development",
     "Navigate to /shared/ or /common/ in the repo. Are reusable components present and documented? "
     "Run SonarQube duplication check: is code duplication below the agreed threshold? "
     "Ask: what is the process when a developer wants to copy an existing function?",
     "all", 1.0),
    ("Development",
     "Open the CI pipeline test results. What is the current pass rate and coverage percentage for core business logic? "
     "Spot-check 5 test files: do assertions test behaviour and outcomes, or just implementation details? "
     "Ask: what is the coverage threshold enforced in CI?",
     "all", 2.0),
    ("Development",
     "Check CONTRIBUTING.md for the branching strategy. "
     "Review Git history: are all changes going through PRs, or are there direct pushes to main? "
     "Verify release tags follow semantic versioning and are present for all releases.",
     "all", 1.0),
    ("Development",
     "Open the CI pipeline config. Is the linter stage set to fail the build on critical violations? "
     "Check the SonarQube quality gate — is it set to 'Passed' as a required check before merge? "
     "Ask: are any linter or SAST suppressions present in the codebase, and are they documented?",
     "all", 2.0),
    ("Development",
     "Pull the SonarQube dead-code report. Any blocker or critical dead-code issues? "
     "Check the feature flag registry — do all flags have expiry dates? "
     "Ask: what is the process for removing a flag after the feature is fully rolled out?",
     "all", 1.0),
    ("Development",
     "Review the codebase for dependency injection patterns. "
     "Are components coded to interfaces or abstract classes, not concrete implementations? "
     "Check that test doubles (mocks/stubs) exist for all external dependencies — can the test suite run without network access?",
     "all", 1.0),
    # ── Testing & Coverage ────────────────────────────────────────────────────
    ("QA",
     "Request the test strategy document. Does it cover all test types (unit, integration, system, UAT, regression, smoke)? "
     "Check that entry and exit criteria are defined for each level. "
     "Ask whether the QA lead and project manager have reviewed and approved it.",
     "all", 2.0),
    ("QA",
     "Open the test management tool (XRAY/TestRail/Zephyr). "
     "Spot-check 5 test cases: do they have descriptive names, expected results, and traceability to a requirement or story? "
     "Check that test data scripts are in /tests/fixtures/ and can be run on a clean environment.",
     "all", 1.0),
    ("QA",
     "Open the CI pipeline config (.github/workflows/, Jenkinsfile, etc.). "
     "Verify automated tests are triggered on every PR to the main integration branch. "
     "Check the last 10 test runs: were there any skipped tests? Ask why.",
     "all", 2.0),
    ("QA",
     "Request the latest coverage report from CI artifacts. "
     "Is the coverage threshold enforced (e.g., --cov-fail-under=80) so that the build fails if it drops? "
     "Ask: which module has the lowest coverage, and what is the plan to address it?",
     "all", 1.0),
    ("QA",
     "Ask for performance test scripts and the last set of results. "
     "Were tests run against realistic data volumes (not just sample data)? "
     "Check that soak or endurance test results from the last 3 months are available and within NFR targets.",
     "all", 2.0),
    ("QA",
     "Compare the staging environment spec to production: same DB version, OS, and external integrations? "
     "Ask: are there any configurations present only in production that cannot be tested in staging? "
     "Flag any production-only config as a test blind spot.",
     "all", 2.0),
    ("QA",
     "Open the defect tracker. Does every defect have a severity, priority, owner, and target fix date? "
     "Check the SLA policy document and compare against the defect age report. "
     "Ask: are there any P1 or P2 defects currently past their SLA target?",
     "all", 2.0),
    # ── DevSecOps ─────────────────────────────────────────────────────────────
    ("DevOps",
     "Open the pipeline definition file in the repo. Verify build, test, security scan, and deploy stages are all present. "
     "Check the last successful run: did all stages pass, including security scans? "
     "Ask: how often does the pipeline fail, and what is the most common failure reason?",
     "all", 2.0),
    ("DevOps",
     "Check the pipeline config for environment-specific variable references. "
     "Verify that .env files with real secrets are absent from the repo (check .gitignore and git history). "
     "Ask: who can approve a deployment to production, and is that access controlled?",
     "all", 2.0),
    ("DevOps",
     "Request the rollback procedure from the deployment runbook. "
     "Ask the DevOps engineer to describe the steps verbally — can they do so confidently? "
     "Check whether a rollback test has been performed in the last 3 months.",
     "all", 2.0),
    ("DevOps",
     "Navigate to /infra/ or /deploy/ in the repo. Are all infrastructure resources defined in IaC? "
     "Check for IaC drift detection: are there any manually provisioned resources not in the IaC state? "
     "Verify the IaC state file is stored remotely (Terraform Cloud, S3, Azure backend) — not locally.",
     "all, cloud", 2.0),
    ("DevOps",
     "Review the PR history for pipeline and IaC changes. Were they peer-reviewed before merge? "
     "Check that there are no manual pipeline configuration changes that are not reflected in IaC. "
     "Ask: who can modify the pipeline definition, and is that access audited?",
     "all", 1.0),
    ("DevOps",
     "Open the CI pipeline config. Identify the SAST, DAST, SCA, and container scan stages. "
     "Check the gating policy: does the pipeline fail on critical findings? "
     "Ask for the last scan reports — any critical findings suppressed without a documented exception?",
     "all", 3.0),
    ("DevOps",
     "Check that pipeline secrets are configured in GitHub Secrets, Azure Key Vault, or Jenkins Credentials. "
     "Review the last 5 pipeline run logs for any secrets visible in plain text. "
     "Ask: is there a secrets rotation policy, and when was the last rotation performed?",
     "all", 3.0),
    ("DevOps",
     "Request the DR runbook and the last restore drill record. "
     "Was the actual RTO within the agreed target? "
     "Check the DR test schedule — is it in the project calendar, not just a one-off event?",
     "all", 3.0),
    ("DevOps",
     "Check the CI/CD tool access matrix. Are service accounts used for deployments, not personal credentials? "
     "Verify the audit log is enabled in the CI/CD tool. "
     "Ask: who has admin access to the CI/CD platform, and has that list been reviewed recently?",
     "all", 2.0),
    # ── Environments & Infrastructure ─────────────────────────────────────────
    ("DevOps",
     "Request the environment comparison matrix. "
     "Check that key configurations (DB version, OS version, external integrations) match between staging and production. "
     "Ask: are there any known configuration differences between environments, and are they documented?",
     "all", 2.0),
    ("DevOps",
     "Navigate to the IaC directory. Are all environment resources defined and version-controlled? "
     "Run or request an IaC drift detection report — are there any manually provisioned resources? "
     "Ask: can a new environment be provisioned from IaC alone, with no manual steps?",
     "all, cloud", 2.0),
    ("DevOps",
     "Review the infrastructure sizing document and auto-scaling policy. "
     "Ask for load test results showing the system under projected peak load. "
     "Check that auto-scaling has been tested — not just configured but actually triggered and verified.",
     "all, cloud", 2.0),
    ("DevOps",
     "Ask for the build failure or environment incident log from the last quarter. "
     "For each incident, is there an RCA document? Are corrective actions tracked in Jira with an owner? "
     "Ask: has the same failure type recurred more than once?",
     "all", 1.0),
    ("DevOps",
     "Request the HA architecture diagram. Does it show multi-AZ or multi-node topology? "
     "Ask for the failover test record: when was the last test, and what was the actual switchover time? "
     "Check that the SLO uptime report reflects the HA configuration.",
     "all, cloud", 3.0),
    ("DevOps",
     "Review the network architecture diagram. "
     "Check that security group rules follow least-privilege: no 0.0.0.0/0 on sensitive ports. "
     "Verify the IaC defines all network resources — no manually configured firewall rules.",
     "all, cloud", 2.0),
    ("DevOps",
     "Confirm that no environment-specific values are hardcoded in source code. "
     "Verify config is injected via environment variables or a secrets vault at runtime. "
     "Check .gitignore excludes .env files and verify they are absent from git history.",
     "all", 3.0),
    # ── Operational Readiness & Reliability ───────────────────────────────────
    ("DevOps",
     "Ask for a sample log trace that spans multiple services. Can a full request be traced end-to-end via a correlation ID? "
     "Check that logs are in structured JSON format. "
     "Verify logs are being ingested into the central logging platform, not just written to local files.",
     "all", 2.0),
    ("DevOps",
     "Open the observability dashboard. Can you see latency P50/P95/P99, error rate, and throughput at a glance? "
     "Check that SLO/SLI targets are visible on the dashboard — not just in a document. "
     "Ask: when an SLO is breached, who is notified, and how quickly?",
     "all", 2.0),
    ("DevOps",
     "Review the alerting rules in the monitoring tool. Are thresholds meaningful, not just default? "
     "Check that every critical alert links to a runbook. "
     "Ask: what is the current on-call rotation, and who is on call right now?",
     "all", 2.0),
    ("DevOps",
     "Request the on-call schedule in PagerDuty or OpsGenie. Does it provide full production-hours coverage with a backup? "
     "Check the escalation matrix document: are contact details current and response SLAs defined per severity level?",
     "all", 2.0),
    ("DevOps",
     "Ask the operations team to locate the runbook for the most recent production incident. "
     "Can they find it without developer help, and does it include steps to diagnose and resolve the issue? "
     "Verify runbooks are in /docs/operations/ — not in a personal notebook or chat tool.",
     "all", 2.0),
    ("DevOps",
     "Request the signed operational readiness checklist. "
     "Check that monitoring, backups, DR, security, and support items are all ticked off. "
     "Ask: were there any open P1/P2 items at go-live? If yes, were they waived and by whom?",
     "all", 3.0),
    ("DevOps",
     "Request post-incident review records for the last 3 incidents. "
     "Check that each includes a timeline, root cause, and follow-up actions with owners. "
     "Ask: are there any recurring incidents that have not had a permanent fix applied?",
     "all", 2.0),
    # ── Compliance & Governance ────────────────────────────────────────────────
    ("Governance",
     "Request the OSS licence scan report. Are there any GPL, AGPL, or other copyleft licences in a proprietary product context? "
     "Check the approved software list from enterprise architecture. "
     "Verify all commercial licences have renewal dates tracked in an inventory.",
     "all", 2.0),
    ("Governance",
     "Ask whether a DPIA or PIA has been completed and signed. "
     "Check the Article 30 record of processing activities. "
     "Verify DPA agreements are in place with all sub-processors who handle personal data.",
     "all, regulated", 3.0),
    ("Governance",
     "Check that audit logs are configured in both the application and infrastructure. "
     "Verify log retention meets compliance requirements (e.g., 1 year minimum). "
     "Ask: are audit logs stored in a tamper-evident system, or could they be deleted by an admin?",
     "all, regulated", 3.0),
    ("Governance",
     "Review the data classification policy. Are PII fields tagged in the data model? "
     "Verify data masking is applied in non-production environments — ask for a sample non-production record. "
     "Check training completion records for data handling training.",
     "all, regulated", 2.0),
    ("Governance",
     "Check the data retention schedule per data class. Is it aligned to legal requirements? "
     "Verify an automated archival or deletion job exists and has been tested. "
     "Ask: what is the legal hold procedure if data deletion must be suspended for litigation?",
     "all, regulated", 2.0),
    ("Governance",
     "Request quarterly access review records for privileged accounts. "
     "Check the segregation of duties matrix — are there any roles that combine approval and execution on critical changes? "
     "Verify that production changes require a formal approval in the ITSM tool.",
     "all", 3.0),
    # ── Performance, Scalability & Resilience ──────────────────────────────────
    ("Development",
     "Request the NFR document. Are SLOs quantified (e.g., P99 < 500ms, availability 99.9%)? "
     "Check for stakeholder sign-off on the NFR document. "
     "Ask: can each NFR be traced to a specific architectural or design decision?",
     "all", 2.0),
    ("Development",
     "Request performance test scripts and the last results report. "
     "Were tests run with realistic data volumes (not sample/seed data)? "
     "Ask for soak test results over a 24-hour period — was the system stable throughout?",
     "all", 2.0),
    ("Development",
     "Ask for APM traces or flame graphs for the last identified bottleneck. "
     "Check Jira performance improvement tickets — are they prioritised and assigned? "
     "Ask: has performance improved over the last 3 sprints, and what is the evidence?",
     "all", 1.0),
    ("Development",
     "Check the architecture for caching layers (Redis, CDN, in-memory). "
     "Verify the cache hit rate is visible in the observability dashboard. "
     "Ask: for which data types is async processing used, and how is failure handled in the async queue?",
     "all", 1.0),
    ("Development",
     "Check the codebase for timeout, retry, and circuit breaker implementations. "
     "Ask for chaos engineering or failure injection test records — has the system been tested under partial failure? "
     "Verify fallback behaviour is documented and tested, not just coded.",
     "all", 2.0),
    ("Development",
     "Request the capacity plan. Does it include projected traffic growth and resource headroom? "
     "Ask for auto-scaling policy test results under simulated peak load. "
     "Verify capacity assumptions were validated in the last quarterly planning cycle.",
     "all, cloud", 2.0),
    # ── APIs, Integrations & Messaging ────────────────────────────────────────
    ("Development",
     "Request the integration register or API dependency map. "
     "Check that each integration has a named owner, SLA, and documented failure mode. "
     "Ask: are there any integrations not in the register that you discovered by reading the code?",
     "all", 2.0),
    ("Development",
     "Ask for the OpenAPI spec and AsyncAPI spec (if applicable). "
     "Spot-check two recently changed APIs: does the spec reflect the current implementation? "
     "Verify the schema registry is in use for Kafka/Avro topics, not informal schema sharing.",
     "all", 2.0),
    ("Development",
     "Check the API gateway configuration in IaC. "
     "Verify rate limiting and authentication are enforced at the gateway, not just in application code. "
     "Ask for the gateway access log — are logs being forwarded to the central logging platform?",
     "all, cloud", 2.0),
    ("Development",
     "Review the integration error handling code. Are timeout values, retry counts, and backoff strategies explicitly configured? "
     "Check that idempotency keys are used for operations that must not be duplicated on retry. "
     "Ask: what happens when an external integration is unavailable — is there a graceful degradation path?",
     "all", 2.0),
    ("Development",
     "Check for a contract test suite (Pact/Spring Cloud Contract) in /tests/contract/. "
     "Are external services stubbed or mocked in integration tests for isolation? "
     "Ask: how do you know when a provider changes their API in a way that breaks your consumer?",
     "all", 1.0),
    ("Development",
     "Request the messaging topology diagram. "
     "Check schema definitions in the schema registry — are breaking changes versioned? "
     "Ask: what is the governance process for schema changes, and who approves them?",
     "all", 2.0),
    # ── User Experience & Accessibility ──────────────────────────────────────
    ("Development",
     "Ask for the user journey map with personas. "
     "Check that E2E tests in /tests/e2e/ cover happy-path and key edge cases per persona. "
     "Ask: have real users or UX researchers tested these journeys, or is this purely developer-tested?",
     "all", 1.0),
    ("Development",
     "Check for a design system or style guide in Storybook or Figma. "
     "Spot-check 3 UI screens: do they match the style guide for typography, spacing, and interaction patterns? "
     "Ask for visual regression test results — are unintended regressions caught automatically?",
     "all", 1.0),
    ("Development",
     "Request the browser compatibility matrix. "
     "Check cross-browser test results for the supported browsers. "
     "Verify responsive design screenshots exist at 320px, 768px, and 1280px breakpoints.",
     "all", 1.0),
    ("Development",
     "Request the WCAG 2.1 AA audit report (axe or Lighthouse). "
     "Check that accessibility issues are tracked in Jira with severity. "
     "Ask: has the application been tested with a screen reader (NVDA, JAWS, or VoiceOver) by a real user or QA?",
     "all", 2.0),
    ("Development",
     "Check whether locale resource files exist in /src/i18n/ or /locales/. "
     "Verify date, time, and currency are formatted using locale-aware libraries, not hardcoded formats. "
     "If RTL support is required, ask for RTL layout test evidence.",
     "all", 1.0),
    ("Development",
     "Review the UI error state designs in Figma or Storybook. "
     "Check that loading indicators appear for operations taking more than 300ms. "
     "Ask: have error states been validated in user testing — or were they designed without user input?",
     "all", 1.0),
    # ── AI Adoption ───────────────────────────────────────────────────────────
    ("Development",
     "Ask each team member to describe which AI tools they use and for what purpose. "
     "Check whether an enterprise or client AI usage policy is in place and has been shared with the team. "
     "Flag any use of public AI tools with proprietary or client data.",
     "all", 2.0),
    ("Development",
     "Ask the team how AI-generated code is reviewed. Is it reviewed with the same rigour as hand-written code? "
     "Check test coverage for AI-generated modules — is it at least 80%? "
     "Verify SAST is applied to AI-generated code before it is merged.",
     "all", 2.0),
    ("Development",
     "Ask to see one Architecture Decision Record where AI was used in analysis. "
     "Verify that AI-assisted design artefacts were reviewed and signed off by the responsible architect. "
     "Check that any AI-generated designs were validated against enterprise or client architecture standards.",
     "all", 1.0),
    ("Development",
     "Ask the team to quantify the productivity impact of AI adoption. "
     "Check sprint velocity data before and after adoption — is the trend improving? "
     "Ask whether the team has any data from retrospectives or surveys on perceived productivity impact.",
     "all", 1.0),
    ("Development",
     "Ask for signed AI usage policy acknowledgements for all team members. "
     "Check that no proprietary or client data is being submitted to public AI tools — how is this enforced? "
     "Verify that the data classification policy is applied to AI prompts, not just to code.",
     "all", 2.0),
]


def style_header_row(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def style_data_rows(ws):
    for i, row in enumerate(ws.iter_rows(min_row=2), 1):
        fill = ROW_FILL_A if i % 2 == 0 else ROW_FILL_B
        for cell in row:
            cell.fill = fill
            cell.font = ROW_FONT
            cell.alignment = ROW_ALIGN


def set_column_widths(ws):
    widths = {
        "A": 26, "B": 12, "C": 16, "D": 42,
        "E": 52, "F": 22, "G": 52, "H": 8, "I": 8,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 28


def write_enriched(src_path: Path, category: str, questions: list, enrichment: list):
    """Read legacy xlsx, write enriched sheet with new schema."""
    print(f"Reading {src_path} …")
    wb_src = openpyxl.load_workbook(src_path)
    ws_src = wb_src.active

    area_col = 1  # 1-indexed
    if category == "Technical":
        question_col = 2
        evidence_col = 3
    else:
        question_col = 2
        evidence_col = 3

    # Extract legacy rows (skip header)
    legacy_rows = []
    current_area = None
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if not any(v for v in row if v is not None):
            continue
        area_val = row[area_col]
        q_val = row[question_col]
        ev_val = row[evidence_col] if len(row) > evidence_col else None

        if area_val and str(area_val).strip():
            current_area = str(area_val).strip()
        if q_val and str(q_val).strip():
            legacy_rows.append((current_area, str(q_val).strip(), str(ev_val).strip() if ev_val else ""))

    if len(legacy_rows) != len(enrichment):
        print(f"  WARNING: {len(legacy_rows)} data rows vs {len(enrichment)} enrichment entries — mismatch!")

    # Build new workbook
    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "Master Template Items"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"

    for (area, question, evidence), (team, guidance, tags, weight) in zip(legacy_rows, enrichment):
        ws.append([area, category, team, question, guidance, tags, evidence, weight, "Yes"])

    style_header_row(ws)
    style_data_rows(ws)
    set_column_widths(ws)

    wb_new.save(src_path)
    print(f"  Saved {len(legacy_rows)} rows to {src_path}")


def main():
    for path in (DELIVERY_PATH, TECHNICAL_PATH):
        if not path.exists():
            sys.exit(f"File not found: {path}")

    write_enriched(DELIVERY_PATH, "Delivery", [], DELIVERY_ENRICHMENT)
    write_enriched(TECHNICAL_PATH, "Technical", [], TECHNICAL_ENRICHMENT)
    print("Done. Both source files updated.")


if __name__ == "__main__":
    main()
