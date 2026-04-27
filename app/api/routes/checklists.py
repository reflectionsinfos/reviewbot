"""
Checklists API Routes
"""
import io
import logging
import os
import tempfile
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional

from app.db.session import get_db
from app.models import Checklist, ChecklistItem, User
from app.api.routes.auth import get_current_user
from app.services.checklist_parser import parse_excel_checklist
from app.services.checklist_service import ChecklistService, _generate_item_code
from app.schemas.checklist import (
    ChecklistItemCreate, ChecklistItemUpdate, ItemReorderReq,
    CloneChecklistResponse, CloneChecklistReq, SyncStrategyReq, SyncResult,
    GlobalChecklistCreate, GlobalChecklistUpdate,
    GlobalChecklistItemCreate, GlobalChecklistItemUpdate,
)

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/")
async def list_checklists(
    is_global: Optional[bool] = None,
    project_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db)
):
    """List checklists with optional filters."""
    return await ChecklistService.list_checklists(db, is_global=is_global, project_id=project_id)


@router.get("/globals/{checklist_id}/area-codes")
async def get_checklist_area_codes(
    checklist_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get area code mapping for a checklist. Admin only.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    return await ChecklistService.get_checklist_area_codes(db, checklist_id)
    
@router.get("/{checklist_id}")
async def get_checklist(
    checklist_id: int,
    include_items: bool = True,
    db: AsyncSession = Depends(get_db)
):
    """Get checklist details"""
    return await ChecklistService.get_checklist(db, checklist_id, include_items)


@router.get("/{checklist_id}/recommendations")
async def get_checklist_recommendations(
    checklist_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get AI-generated recommendations for checklist improvement"""
    return await ChecklistService.get_checklist_recommendations(db, checklist_id)


@router.post("/{checklist_id}/optimize")
async def optimize_checklist(
    checklist_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Generate AI recommendations for checklist optimization"""
    return await ChecklistService.optimize_checklist(db, checklist_id)


@router.get("/templates/global")
async def get_global_checklist_templates(
    type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get global checklist templates visible to the current user's organization."""
    return await ChecklistService.get_global_checklist_templates(
        db, type, organization_id=current_user.organization_id
    )


@router.get("/templates/download-excel-template")
async def download_excel_template(
    type: Optional[str] = Query(None, description="Checklist type: 'delivery', 'technical', or 'master'"),
):
    """
    Download a pre-filled Excel checklist template (.xlsx).

    - type=master (or omitted) → single sheet "Master Template Items" with all 156 items
    - type=delivery  → single sheet "Master Template Items" with 49 delivery items only
    - type=technical → single sheet "Master Template Items" with 107 technical items only

    Column layout: Area | Category | Team Category | Question | Guidance |
                   Applicability Tags | Evidence | Weight | Review?
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from pathlib import Path

    # ── column layout ────────────────────────────────────────────────────────
    HEADERS = [
        "Area", "Category", "Team Category", "Question", "Guidance",
        "Applicability Tags", "Evidence", "Weight", "Review?",
    ]
    COL_WIDTHS = [26, 12, 16, 42, 52, 22, 52, 8, 8]

    # ── styling ──────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
    ALT_FILL  = PatternFill("solid", fgColor="EEF2F7")
    THIN      = Side(style="thin", color="BDC8D6")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    DATA_FONT = Font(name="Calibri", size=10)

    def _style_header(ws):
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = BORDER

    def _style_data_row(ws, row_idx: int, alternate: bool):
        for cell in ws[row_idx]:
            if alternate:
                cell.fill = ALT_FILL
            cell.font      = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = BORDER
        ws.row_dimensions[row_idx].height = 36

    def _set_col_widths(ws):
        for i, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _add_data_validation(ws, max_row: int):
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        dv.sqref = f"I2:I{max(max_row + 200, 300)}"
        ws.add_data_validation(dv)

    def _populate_sheet(ws, rows):
        """Write rows to the worksheet."""
        ws.append(HEADERS)
        _style_header(ws)
        for idx, row_data in enumerate(rows, 1):
            ws.append(row_data)
            _style_data_row(ws, ws.max_row, alternate=(idx % 2 == 0))
        _set_col_widths(ws)
        _add_data_validation(ws, ws.max_row)
        ws.freeze_panes = "A2"

    # ── locate source files ───────────────────────────────────────────────────
    tmpl_root = Path(__file__).resolve().parents[3] / "data" / "templates"
    candidates = [tmpl_root / "reviewbot" / "v1", tmpl_root]
    base = next((p for p in candidates if (p / "standard-delivery.xlsx").exists()), None)

    def _read_source(src_path: Path) -> list:
        """Read enriched source file (new unified column schema)."""
        try:
            df = pd.read_excel(src_path, sheet_name="Master Template Items")
        except Exception:
            return []
        rows, current_area = [], ""
        for _, row in df.iterrows():
            av = row.get("Area")
            if pd.notna(av) and str(av).strip():
                current_area = str(av).strip()
            q = row.get("Question")
            if pd.isna(q) or not str(q).strip():
                continue

            def _s(col):
                v = row.get(col)
                return str(v).strip() if pd.notna(v) and str(v).strip() else ""

            rows.append([
                current_area,
                _s("Category"),
                _s("Team Category"),
                str(q).strip(),
                _s("Guidance"),
                _s("Applicability Tags"),
                _s("Evidence"),
                float(row.get("Weight") or 1.0),
                "Yes",
            ])
        return rows

    def _read_delivery():
        return _read_source(base / "standard-delivery.xlsx") if base else []

    def _read_technical():
        return _read_source(base / "standard-technical.xlsx") if base else []

    # ── build workbook ───────────────────────────────────────────────────────
    wb = Workbook()
    checklist_type = (type or "master").strip().lower()

    if checklist_type == "delivery":
        ws = wb.active
        ws.title = "Master Template Items"
        _populate_sheet(ws, _read_delivery())
        fname = "reviewbot_delivery_template.xlsx"

    elif checklist_type == "technical":
        ws = wb.active
        ws.title = "Master Template Items"
        _populate_sheet(ws, _read_technical())
        fname = "reviewbot_technical_template.xlsx"

    else:
        # master (default) → all 156 items in one sheet
        ws = wb.active
        ws.title = "Master Template Items"
        _populate_sheet(ws, _read_delivery() + _read_technical())
        fname = "reviewbot_master_template.xlsx"

    # ── stream ───────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/templates/use/{template_id}")
async def use_global_template(
    template_id: int,
    project_id: int,
):
    """Deprecated — use POST /api/checklists/{template_id}/clone-to-project/{project_id} instead."""
    raise HTTPException(
        status_code=410,
        detail=(
            "This endpoint is deprecated. Use "
            f"POST /api/checklists/{template_id}/clone-to-project/<project_id> instead. "
            "The new endpoint requires authentication and tracks the source template for sync."
        )
    )


@router.post("/{checklist_id}/clone-to-project/{project_id}", response_model=CloneChecklistResponse)
async def clone_checklist_to_project(
    checklist_id: int,
    project_id: int,
    req: CloneChecklistReq = CloneChecklistReq(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Clone a global checklist into a project-specific checklist"""
    try:
        new_checklist = await ChecklistService.clone_checklist_to_project(
            db, checklist_id, project_id, req, current_user
        )
        
        return CloneChecklistResponse(
            id=new_checklist.id,
            name=new_checklist.name,
            type=new_checklist.type,
            version=new_checklist.version,
            project_id=new_checklist.project_id,
            is_global=new_checklist.is_global,
            item_count=len(new_checklist.items),
            source_checklist_id=new_checklist.source_checklist_id
        )
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{checklist_id}/sync-from-global", response_model=SyncResult)
async def sync_from_global(
    checklist_id: int,
    req: SyncStrategyReq,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Sync a project checklist with its original global template"""
    try:
        sync_data = await ChecklistService.sync_from_global(db, checklist_id, req, current_user)
        return SyncResult(
            added=sync_data["added"],
            updated=sync_data["updated"],
            flagged_removed=sync_data["flagged_removed"],
            strategy_used=sync_data["strategy_used"],
            flagged_items=sync_data["flagged_items"]
        )
        
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{checklist_id}/items")
async def add_checklist_item(
    checklist_id: int,
    item_in: ChecklistItemCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Add item to a project-specific checklist"""
    try:
        new_item = await ChecklistService.add_checklist_item(db, checklist_id, item_in, current_user)

        return {
            "id": new_item.id,
            "checklist_id": new_item.checklist_id,
            "area": new_item.area,
            "question": new_item.question,
            "category": new_item.category,
            "weight": new_item.weight,
            "is_review_mandatory": new_item.is_review_mandatory,
            "expected_evidence": new_item.expected_evidence,
            "team_category": new_item.team_category,
            "guidance": new_item.guidance,
            "applicability_tags": new_item.applicability_tags,
            "item_code": new_item.item_code,
            "order": new_item.order
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{checklist_id}/items/{item_id}")
async def update_checklist_item(
    checklist_id: int,
    item_id: int,
    item_in: ChecklistItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update item on a project-specific checklist"""
    try:
        item = await ChecklistService.update_checklist_item(db, checklist_id, item_id, item_in, current_user)
        
        return {
            "id": item.id,
            "checklist_id": item.checklist_id,
            "area": item.area,
            "question": item.question,
            "category": item.category,
            "weight": item.weight,
            "is_review_mandatory": item.is_review_mandatory,
            "expected_evidence": item.expected_evidence,
            "team_category": item.team_category,
            "guidance": item.guidance,
            "applicability_tags": item.applicability_tags,
            "item_code": item.item_code,
            "order": item.order
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{checklist_id}/items/{item_id}")
async def delete_checklist_item(
    checklist_id: int,
    item_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete item from a project-specific checklist"""
    try:
        await ChecklistService.delete_checklist_item(db, checklist_id, item_id, current_user)
        return {"message": "Item deleted successfully"}
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{checklist_id}/items/reorder")
async def reorder_checklist_items(
    checklist_id: int,
    orders: List[ItemReorderReq],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Reorder multiple items on a project-specific checklist"""
    try:
        await ChecklistService.reorder_checklist_items(db, checklist_id, orders, current_user)
        return {"message": "Items reordered successfully"}
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{checklist_id}")
async def delete_checklist(
    checklist_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Delete a project-specific checklist and all its items.
    Blocked if any reviews (manual or autonomous) have been run against this checklist.
    Global checklists cannot be deleted via this endpoint.
    """
    try:
        await ChecklistService.delete_checklist(db, checklist_id, current_user)
        return {"message": f"Checklist deleted successfully"}

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error deleting checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────────────────────────────────────
# Global Checklist Management (Admin Only)
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/globals")
async def create_global_checklist(
    req: GlobalChecklistCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Create a new global checklist (empty, no items).
    Requires admin role.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    try:
        checklist = await ChecklistService.create_global_checklist(db, req)
        logger.info(f"Global checklist created: {checklist.name} (ID: {checklist.id}) by {current_user.email}")

        return {
            "id": checklist.id,
            "name": checklist.name,
            "type": checklist.type,
            "version": checklist.version,
            "is_global": True,
            "organization_id": checklist.organization_id,
            "item_count": 0
        }

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error creating global checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/globals/{checklist_id}")
async def update_global_checklist(
    checklist_id: int,
    req: GlobalChecklistUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Rename a global checklist or update its version.
    Requires admin role.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    try:
        checklist = await ChecklistService.update_global_checklist(db, checklist_id, req)
        logger.info(f"Global checklist updated: {checklist.name} (ID: {checklist.id}) by {current_user.email}")

        return {
            "id": checklist.id,
            "name": checklist.name,
            "version": checklist.version
        }

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error updating global checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/globals/{checklist_id}")
async def delete_global_checklist(
    checklist_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Delete a global checklist.
    Requires admin role.
    Blocked if any project checklists are cloned from it or if reviews exist.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    try:
        await ChecklistService.delete_global_checklist(db, checklist_id)
        logger.info(f"Global checklist (ID: {checklist_id}) deleted by {current_user.email}")
        return {"message": f"Global checklist deleted successfully"}

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error deleting global checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/globals/upload")
async def upload_global_checklist(
    file: UploadFile = File(..., description="Excel checklist file (.xlsx)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Upload an Excel checklist file and create one or more global templates.
    Parses sheets named 'Delivery Check List V 1.0' and/or 'Technical Check List V 1.0'.
    Requires admin role.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    # Validate file type
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xlsm files are supported")

    # Validate file size (25 MB limit)
    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 25 MB)")

    # Save to temp file and parse
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        parsed = await parse_excel_checklist(tmp_path)
        checklists_data = parsed.get("checklists", {})

        if not checklists_data:
            raise HTTPException(status_code=422, detail="No recognisable checklist sheets found in this file")

        created = []
        for checklist_type, items in checklists_data.items():
            if not items:
                continue

            # Derive a name from the original filename (strip extension)
            base_name = os.path.splitext(filename)[0]
            checklist_name = f"{base_name} ({checklist_type.capitalize()})"

            checklist = Checklist(
                name=checklist_name,
                type=checklist_type,
                version="1.0",
                is_global=True,
                project_id=None,
                organization_id=current_user.organization_id,
            )
            db.add(checklist)
            await db.flush()  # get checklist.id

            for order, item_data in enumerate(items):
                area = item_data.get("area", "General")
                # Generate AREA-NNN code (registers area abbreviation on checklist.area_codes).
                # Must flush after each add so the next call sees this item when computing max_seq.
                item_code = await _generate_item_code(db, checklist.id, area)
                item = ChecklistItem(
                    checklist_id=checklist.id,
                    item_code=item_code,
                    area=area,
                    question=item_data.get("question", ""),
                    category=item_data.get("category", checklist_type),
                    weight=item_data.get("weight", 1.0),
                    is_review_mandatory=item_data.get("is_review_mandatory", True),
                    expected_evidence=item_data.get("expected_evidence"),
                    team_category=item_data.get("team_category"),
                    guidance=item_data.get("guidance"),
                    applicability_tags=item_data.get("applicability_tags"),
                    order=order,
                )
                db.add(item)
                await db.flush()  # make this item visible to next _generate_item_code query
            created.append({
                "id": checklist.id,
                "name": checklist.name,
                "type": checklist.type,
                "version": checklist.version,
                "item_count": len(items),
            })

        await db.commit()
        logger.info(
            f"Global checklist(s) uploaded from '{filename}' by {current_user.email}: "
            f"{[c['name'] for c in created]}"
        )
        return {"created": created}

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error uploading global checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to parse Excel file: {str(e)}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/globals/{checklist_id}/upload-items")
async def upload_items_to_global_checklist(
    checklist_id: int,
    file: UploadFile = File(..., description="Excel checklist file (.xlsx)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Replace all items on an existing global checklist by uploading an Excel file.

    The file must use the unified column layout:
      Area | Team Category | Question | Guidance | Applicability Tags | Evidence | Weight | Review?

    The correct sheet is selected automatically based on the checklist type:
      - delivery  → reads "Delivery Check List V 1.0" sheet
      - technical → reads "Technical Check List V 1.0" sheet

    All existing items are deleted and replaced with the uploaded content.
    Requires admin role.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xlsm files are supported")

    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 25 MB)")

    tmp_path = None
    try:
        # Resolve the checklist and confirm it exists and is global
        from sqlalchemy import select as sa_select, delete as sa_delete
        result = await db.execute(
            sa_select(Checklist).where(Checklist.id == checklist_id, Checklist.is_global == True)
        )
        checklist = result.scalar_one_or_none()
        if checklist is None:
            raise HTTPException(status_code=404, detail="Global checklist not found")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        parsed = await parse_excel_checklist(tmp_path)
        checklists_data = parsed.get("checklists", {})

        # Pick items matching this checklist's type
        checklist_type = checklist.type  # "delivery", "technical", or "master"
        items_data = checklists_data.get(checklist_type)
        if not items_data:
            if checklist_type == "master":
                # Master sheet parsed via parse_master_sheet — key is "master"
                items_data = checklists_data.get("master")
            if not items_data:
                # Fallback: accept any single sheet present in the file
                if len(checklists_data) == 1:
                    items_data = next(iter(checklists_data.values()))
                else:
                    raise HTTPException(
                        status_code=422,
                        detail=(
                            f"No '{checklist_type}' sheet found in the uploaded file. "
                            "For master templates upload a file with a 'Master Template Items' sheet. "
                            "For delivery/technical templates use the matching sheet name."
                        ),
                    )

        if not items_data:
            raise HTTPException(status_code=422, detail="The uploaded file contains no checklist items")

        # Delete all existing items for this checklist
        await db.execute(
            sa_delete(ChecklistItem).where(ChecklistItem.checklist_id == checklist_id)
        )
        # Reset area codes so they are regenerated cleanly
        checklist.area_codes = {}
        await db.flush()

        # Insert new items
        for order, item_data in enumerate(items_data):
            area = item_data.get("area", "General")
            item_code = await _generate_item_code(db, checklist.id, area)
            item = ChecklistItem(
                checklist_id=checklist.id,
                item_code=item_code,
                area=area,
                question=item_data.get("question", ""),
                category=item_data.get("category", checklist_type),
                weight=item_data.get("weight", 1.0),
                is_review_mandatory=item_data.get("is_review_mandatory", True),
                expected_evidence=item_data.get("expected_evidence"),
                team_category=item_data.get("team_category"),
                guidance=item_data.get("guidance"),
                applicability_tags=item_data.get("applicability_tags"),
                order=order,
            )
            db.add(item)
            await db.flush()

        await db.commit()
        logger.info(
            f"Items replaced on global checklist {checklist_id} ('{checklist.name}') "
            f"from '{filename}' by {current_user.email}: {len(items_data)} items"
        )
        return {
            "checklist_id": checklist_id,
            "checklist_name": checklist.name,
            "items_replaced": len(items_data),
        }

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error uploading items to global checklist {checklist_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to parse Excel file: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/globals/{checklist_id}/items")
async def add_item_to_global_checklist(
    checklist_id: int,
    item_in: GlobalChecklistItemCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Add an item to a global checklist.
    Requires admin role.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    try:
        item = await ChecklistService.add_item_to_global_checklist(db, checklist_id, item_in)
        logger.info(
            f"Item added to global checklist: checklist_id={checklist_id}, "
            f"item_code={item.item_code} by {current_user.email}"
        )

        return {
            "id": item.id,
            "checklist_id": item.checklist_id,
            "item_code": item.item_code,
            "area": item.area,
            "question": item.question,
            "category": item.category,
            "weight": item.weight,
            "is_review_mandatory": item.is_review_mandatory,
            "expected_evidence": item.expected_evidence,
            "team_category": item.team_category,
            "guidance": item.guidance,
            "applicability_tags": item.applicability_tags,
            "order": item.order
        }

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error adding item to global checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/globals/{checklist_id}/items/{item_id}")
async def update_item_in_global_checklist(
    checklist_id: int,
    item_id: int,
    item_in: GlobalChecklistItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Update an item on a global checklist.
    Requires admin role.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    try:
        item = await ChecklistService.update_item_in_global_checklist(db, checklist_id, item_id, item_in)
        logger.info(
            f"Item updated in global checklist: checklist_id={checklist_id}, item_id={item_id} by {current_user.email}"
        )

        return {
            "id": item.id,
            "checklist_id": item.checklist_id,
            "item_code": item.item_code,
            "area": item.area,
            "question": item.question,
            "category": item.category,
            "weight": item.weight,
            "is_review_mandatory": item.is_review_mandatory,
            "expected_evidence": item.expected_evidence,
            "team_category": item.team_category,
            "guidance": item.guidance,
            "applicability_tags": item.applicability_tags,
            "order": item.order
        }

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error updating item in global checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/globals/{checklist_id}/items/{item_id}")
async def delete_item_from_global_checklist(
    checklist_id: int,
    item_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Delete an item from a global checklist.
    Requires admin role.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    try:
        await ChecklistService.delete_item_from_global_checklist(db, checklist_id, item_id)
        logger.info(
            f"Item deleted from global checklist: checklist_id={checklist_id}, item_id={item_id} by {current_user.email}"
        )

        return {"message": "Item deleted"}

    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"Error deleting item from global checklist: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
