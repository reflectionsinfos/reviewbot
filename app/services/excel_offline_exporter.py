"""
Generates a response-ready Excel workbook for offline reviews.

The reviewer downloads this file, fills in the Response/Comments columns,
then uploads it back to the tool.
"""
import io
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Colour palette
_BLUE_HEADER = "1E3A5F"
_GREY_LOCKED = "F2F2F2"
_WHITE_EDITABLE = "FFFFFF"
_AMBER = "FFC000"
_GREEN = "70AD47"
_RED = "FF0000"
_LIGHT_BLUE_ROW = "EBF3FB"

_THIN = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def generate_offline_excel(
    project_name: str,
    checklist_name: str,
    reviewer_name: str,
    items: list,
    due_date: Optional[datetime] = None,
    admin_message: Optional[str] = None,
) -> bytes:
    """
    Build an Excel workbook for an offline review.

    Args:
        project_name: Name of the project being reviewed.
        checklist_name: Name of the checklist template.
        reviewer_name: Full name of the assigned reviewer.
        items: List of ChecklistItem ORM objects.
        due_date: Optional target completion date.
        admin_message: Optional message from the admin to include.

    Returns:
        Raw bytes of the .xlsx workbook (suitable for HTTP streaming or email attach).
    """
    wb = Workbook()

    _build_instructions_sheet(wb, project_name, checklist_name, reviewer_name, due_date, admin_message)
    _build_response_sheet(wb, project_name, checklist_name, items)

    # Remove default empty sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet 1 — Instructions
# ---------------------------------------------------------------------------

def _build_instructions_sheet(
    wb: Workbook,
    project_name: str,
    checklist_name: str,
    reviewer_name: str,
    due_date: Optional[datetime],
    admin_message: Optional[str],
) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    header_fill = PatternFill("solid", fgColor=_BLUE_HEADER)
    header_font = Font(bold=True, color="FFFFFF", size=14)
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    section_font = Font(bold=True, size=12, color=_BLUE_HEADER)

    # Title row
    ws.merge_cells("A1:B1")
    ws["A1"] = "ReviewBot — Offline Review Checklist"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Metadata
    meta = [
        ("Project", project_name),
        ("Checklist", checklist_name),
        ("Reviewer", reviewer_name),
        ("Sent Date", datetime.utcnow().strftime("%d %b %Y")),
        ("Due Date", due_date.strftime("%d %b %Y") if due_date else "—"),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value).font = value_font
        ws.row_dimensions[i].height = 20

    if admin_message:
        ws.cell(row=9, column=1, value="Message from Admin").font = label_font
        ws.cell(row=10, column=1, value=admin_message).font = value_font
        ws.merge_cells("A10:B10")
        ws["A10"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[10].height = 60
        section_start = 12
    else:
        section_start = 10

    # Instructions section
    ws.cell(row=section_start, column=1, value="How to complete this review").font = section_font
    ws.merge_cells(f"A{section_start}:B{section_start}")

    steps = [
        "1. Go to the 'ReviewBot Response' sheet (tab at the bottom of this file).",
        "2. For each checklist item, fill in the Response column using the dropdown: Yes / No / Partial / N/A.",
        "3. Add your Comments explaining your response or noting any gaps.",
        "4. Optionally add Evidence Links (URLs or document references, comma-separated).",
        "5. Save the file and upload it back using the link in the invitation email.",
        "   • Leave the Code, Area, Team, Question, and Evidence columns unchanged.",
    ]
    for j, step in enumerate(steps, start=section_start + 1):
        cell = ws.cell(row=j, column=1, value=step)
        cell.font = value_font
        ws.merge_cells(f"A{j}:B{j}")
        ws.row_dimensions[j].height = 18

    # RAG legend
    rag_row = section_start + len(steps) + 2
    ws.cell(row=rag_row, column=1, value="Response guide").font = section_font
    ws.merge_cells(f"A{rag_row}:B{rag_row}")

    rag = [
        ("Yes", _GREEN, "Fully in place — evidence available"),
        ("Partial", _AMBER, "Partially done — gaps noted in Comments"),
        ("No", "FF0000", "Not in place — explain in Comments"),
        ("N/A", "AAAAAA", "Not applicable to this project"),
    ]
    for k, (resp, colour, meaning) in enumerate(rag, start=rag_row + 1):
        cell_resp = ws.cell(row=k, column=1, value=resp)
        cell_resp.font = Font(bold=True, color="FFFFFF")
        cell_resp.fill = PatternFill("solid", fgColor=colour)
        cell_resp.alignment = Alignment(horizontal="center")
        ws.cell(row=k, column=2, value=meaning).font = value_font
        ws.row_dimensions[k].height = 20


# ---------------------------------------------------------------------------
# Sheet 2 — Response items
# ---------------------------------------------------------------------------

def _build_response_sheet(
    wb: Workbook,
    project_name: str,
    checklist_name: str,
    items: list,
) -> None:
    ws = wb.create_sheet("ReviewBot Response", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "G3"  # Freeze locked columns + header row

    # Column layout: (header, width, locked)
    columns = [
        ("Code", 10, True),
        ("Area", 22, True),
        ("Team", 12, True),
        ("Question", 50, True),
        ("Expected Evidence", 30, True),
        ("Guidance", 30, True),
        ("Response ▼", 14, False),
        ("Comments", 40, False),
        ("Evidence Links", 35, False),
    ]

    header_fill = PatternFill("solid", fgColor=_BLUE_HEADER)
    locked_fill = PatternFill("solid", fgColor=_GREY_LOCKED)
    edit_fill = PatternFill("solid", fgColor=_WHITE_EDITABLE)
    alt_fill = PatternFill("solid", fgColor=_LIGHT_BLUE_ROW)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    locked_font = Font(size=10, color="333333")
    edit_font = Font(size=10)

    # Row 1: project/checklist info bar
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Project: {project_name}   |   Checklist: {checklist_name}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=10)
    ws["A1"].fill = PatternFill("solid", fgColor="2E75B6")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 20

    # Row 2: column headers
    for col_idx, (header, width, _locked) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws.row_dimensions[2].height = 28

    # Data validation dropdown for Response column (col 7)
    dv = DataValidation(
        type="list",
        formula1='"Yes,No,Partial,N/A"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid response",
        error="Please select Yes, No, Partial, or N/A from the dropdown.",
    )
    ws.add_data_validation(dv)

    # Data rows
    for row_offset, item in enumerate(items):
        row = row_offset + 3
        use_alt = row_offset % 2 == 0
        row_fill = alt_fill if use_alt else edit_fill

        values = [
            getattr(item, "item_code", "") or "",
            getattr(item, "area", "") or "",
            getattr(item, "team_category", "") or "",
            getattr(item, "question", "") or "",
            getattr(item, "expected_evidence", "") or "",
            getattr(item, "guidance", "") or "",
            "",  # Response (editable)
            "",  # Comments (editable)
            "",  # Evidence Links (editable)
        ]

        for col_idx, (value, (_, _, locked)) in enumerate(zip(values, columns), start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if locked:
                cell.font = locked_font
                cell.fill = locked_fill if not use_alt else PatternFill("solid", fgColor="E8E8E8")
            else:
                cell.font = edit_font
                cell.fill = row_fill

        # Attach dropdown validation to Response cell
        response_cell = ws.cell(row=row, column=7)
        dv.add(response_cell)

        ws.row_dimensions[row].height = 40

    # Protect locked columns (sheets-level protection)
    # Mark editable columns as not locked in their protection attribute
    ws.protection.sheet = True
    ws.protection.password = "reviewbot"
    ws.protection.enable()

    for row_offset in range(len(items)):
        row = row_offset + 3
        for col_idx in (7, 8, 9):  # Response, Comments, Evidence Links
            cell = ws.cell(row=row, column=col_idx)
            cell.protection = _unlocked()


def _unlocked():
    from openpyxl.styles.protection import Protection
    return Protection(locked=False)
